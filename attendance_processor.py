import argparse
import os
from datetime import datetime, date, time, timedelta
from typing import List, Dict, Any, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

EMPLOYEE_MARKER = "Employee ID:"


def extract_employees_from_file(file_path: str) -> List[Dict[str, str]]:
    """
    استخراج بيانات الموظفين من ملف Excel بدون معالجة الحضور - محسن للسرعة
    يُستخدم للمزامنة السريعة مع قاعدة البيانات
    """
    try:
        print(f"🔍 استخراج بيانات الموظفين من: {file_path}")
        
        wb = load_workbook(file_path, data_only=True, read_only=True)
        ws = wb.worksheets[0]  # استخدام الورقة الأولى
        
        employees = []
        employees_found = 0
        max_rows = min(ws.max_row, 2000)  # تحديد الحد الأقصى لتجنب الملفات الكبيرة جداً
        
        print(f"📊 فحص سريع للموظفين في {max_rows} صف...")
        
        # استخدام iter_rows للحصول على أداء أفضل
        for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, min_col=1, max_col=1, values_only=True), 1):
            cell_value = row[0]
            
            if cell_value and EMPLOYEE_MARKER in str(cell_value):
                # استخراج بيانات الموظف
                employee_data = parse_employee_line(str(cell_value))
                if employee_data:
                    employees.append(employee_data)
                    employees_found += 1
                    
                    # طباعة التقدم كل 10 موظفين لتجنب الإفراط في الطباعة
                    if employees_found % 10 == 0:
                        print(f"👥 تم العثور على {employees_found} موظف...")
                    elif employees_found <= 5:  # طباعة أول 5 موظفين فقط
                        print(f"👤 موظف #{employees_found}: {employee_data['EmployeeID']} - {employee_data['Name']}")
            
            # إيقاف مبكر إذا وجدنا عدد كبير من الموظفين (تحسين الأداء)
            if employees_found >= 100:  # حد أقصى معقول للموظفين
                print(f"⚡ تم الوصول للحد الأقصى ({employees_found} موظف) - إيقاف البحث")
                break
        
        wb.close()
        print(f"✅ تم استخراج {len(employees)} موظف بنجاح")
        return employees
        
    except Exception as e:
        print(f"❌ خطأ في استخراج بيانات الموظفين: {e}")
        return []


def parse_employee_line(line: str) -> Optional[Dict[str, str]]:
    """
    تحليل سطر بيانات الموظف واستخراج المعلومات
    مثال: "Employee ID: 102,First Name: Ali,Department: Driver support"
    """
    try:
        parts = line.split(',')
        employee_data = {}
        
        for part in parts:
            part = part.strip()
            if ':' in part:
                key, value = part.split(':', 1)
                key = key.strip()
                value = value.strip()
                
                if key == "Employee ID":
                    employee_data['EmployeeID'] = value
                elif key in ["First Name", "Name"]:
                    employee_data['Name'] = value
                elif key == "Department":
                    employee_data['Department'] = value
        
        # التحقق من وجود البيانات الأساسية
        if 'EmployeeID' in employee_data and 'Name' in employee_data:
            return employee_data
        
        return None
        
    except Exception as e:
        print(f"❌ خطأ في تحليل سطر الموظف: {e}")
        return None


def get_all_active_requests(start_date: date = None, end_date: date = None) -> Dict[str, Dict[str, Any]]:
    """
    جلب جميع الطلبات النشطة من Firebase مرة واحدة وتنظيمها حسب employeeId
    هذا أكثر كفاءة من البحث عن كل موظف على حدة
    """
    try:
        # محاولة استيراد Firebase
        try:
            from firebase_config import get_db
            db = get_db()
            print(f"✅ تم الاتصال بـ Firebase بنجاح")
        except ImportError as e:
            print(f"⚠️ Firebase غير متاح: {e}")
            return {}
        except Exception as e:
            print(f"⚠️ خطأ في الاتصال بـ Firebase: {e}")
            return {}
        
        if not db:
            print("⚠️ Firebase غير متصل، سيتم تجاهل الطلبات")
            return {}
        
        print(f"🔍 جلب جميع الطلبات النشطة من قاعدة البيانات...")
        
        # جلب جميع الطلبات النشطة مرة واحدة
        requests_ref = db.collection('requests')
        query = requests_ref.where('status', '==', 'active')
        
        docs = list(query.stream())
        print(f"📋 تم العثور على {len(docs)} طلب نشط في قاعدة البيانات")
        
        # تنظيم الطلبات حسب employeeId
        employee_requests = {}
        
        for doc in docs:
            data = doc.to_dict()
            employee_id = str(data.get('employeeId', ''))
            
            if not employee_id:
                continue
                
            request_date_str = data.get('reqDate')
            if not request_date_str:
                continue
                
            try:
                # تحويل التاريخ من string إلى date
                if isinstance(request_date_str, str):
                    request_date = datetime.strptime(request_date_str, '%Y-%m-%d').date()
                else:
                    request_date = request_date_str
                    
                # فلترة التواريخ إذا تم تحديدها
                if start_date and end_date:
                    if request_date < start_date or request_date > end_date:
                        continue
                        
                # إنشاء entry للموظف إذا لم يكن موجود
                if employee_id not in employee_requests:
                    employee_requests[employee_id] = {
                        'overtime_requests': [],
                        'leave_requests': [],
                        'overtime_dates': set(),
                        'leave_dates': set()
                    }
                
                request_type = data.get('kind', '')
                request_reason = data.get('reason', '')
                
                request_info = {
                    'date': request_date,
                    'reason': request_reason,
                    'id': data.get('id'),
                    'supervisor': data.get('supervisor', '')
                }
                
                if request_type == 'overtime':
                    employee_requests[employee_id]['overtime_requests'].append(request_info)
                    employee_requests[employee_id]['overtime_dates'].add(request_date)
                    
                elif request_type == 'leave':
                    employee_requests[employee_id]['leave_requests'].append(request_info)
                    employee_requests[employee_id]['leave_dates'].add(request_date)
                    
            except ValueError as e:
                print(f"   ❌ خطأ في تحويل التاريخ {request_date_str}: {e}")
                continue
        
        # إحصائيات مفيدة
        total_overtime_requests = sum(len(emp_data['overtime_requests']) for emp_data in employee_requests.values())
        total_leave_requests = sum(len(emp_data['leave_requests']) for emp_data in employee_requests.values())
        
        print(f"📊 تم تنظيم الطلبات:")
        print(f"   - موظفين لديهم طلبات: {len(employee_requests)}")
        print(f"   - إجمالي طلبات الإضافي: {total_overtime_requests}")
        print(f"   - إجمالي طلبات الإجازة: {total_leave_requests}")
        
        # عرض مثال للموظف 102 إذا كان موجود
        if '102' in employee_requests:
            emp_102 = employee_requests['102']
            print(f"🔍 مثال - الموظف 102:")
            print(f"   - طلبات إضافي: {len(emp_102['overtime_requests'])}")
            print(f"   - طلبات إجازة: {len(emp_102['leave_requests'])}")
            if emp_102['overtime_requests']:
                req = emp_102['overtime_requests'][0]
                print(f"   - مثال طلب إضافي: {req['date']} - {req['reason']}")
        
        return employee_requests
        
    except Exception as e:
        print(f"❌ خطأ في جلب الطلبات: {e}")
        import traceback
        traceback.print_exc()
        return {}


def process_employee_requests_from_cache(employee_id: str, daily_data: List[Dict] = None, all_requests: Dict[str, Dict[str, Any]] = None) -> Dict[str, Any]:
    """
    حساب طلبات موظف معين من البيانات المُحملة مسبقاً
    هذا يتجنب الحاجة لاستعلام Firebase لكل موظف على حدة
    """
    if not all_requests or str(employee_id) not in all_requests:
        return {"overtime_hours": 0, "leave_days": 0, "overtime_requests": [], "leave_requests": [], "overtime_dates": [], "leave_dates": []}
    
    employee_data = all_requests[str(employee_id)]
    all_overtime_requests = employee_data.get('overtime_requests', [])
    all_leave_requests = employee_data.get('leave_requests', [])
    
    # تحديد فترة البيانات اليومية للموظف
    start_date = None
    end_date = None
    if daily_data:
        dates = [d.get("Date") for d in daily_data if d.get("Date")]
        if dates:
            start_date = min(dates)
            end_date = max(dates)
    
    # فلترة الطلبات حسب فترة البيانات اليومية
    overtime_requests = []
    leave_requests = []
    overtime_dates = set()
    leave_dates = set()
    
    # فلترة طلبات الإضافي
    for req in all_overtime_requests:
        req_date = req['date']
        if start_date and end_date:
            if start_date <= req_date <= end_date:
                overtime_requests.append(req)
                overtime_dates.add(req_date)
        else:
            overtime_requests.append(req)
            overtime_dates.add(req_date)
    
    # فلترة طلبات الإجازة
    for req in all_leave_requests:
        req_date = req['date']
        if start_date and end_date:
            if start_date <= req_date <= end_date:
                leave_requests.append(req)
                leave_dates.add(req_date)
        else:
            leave_requests.append(req)
            leave_dates.add(req_date)
    
    # حساب الساعات الإضافية الفعلية من البيانات اليومية
    overtime_hours = 0
    if daily_data and overtime_dates:
        for day_data in daily_data:
            day_date = day_data.get("Date")
            if day_date in overtime_dates:
                day_hours = day_data.get("DayHours", 0)
                if day_hours > 7:
                    additional_hours = day_hours - 7
                    overtime_hours += additional_hours
                    print(f"   📅 {day_date}: {day_hours} ساعة، إضافي: {additional_hours} ساعة")
    
    # حساب أيام الإجازة
    leave_days = len(leave_dates)
    
    if overtime_hours > 0 or leave_days > 0:
        print(f"📊 موظف {employee_id}: {overtime_hours} ساعة إضافية، {leave_days} يوم إجازة")
    
    return {
        "overtime_hours": round(overtime_hours, 2),
        "leave_days": leave_days,
        "overtime_requests": overtime_requests,
        "leave_requests": leave_requests,
        "overtime_dates": list(overtime_dates),
        "leave_dates": list(leave_dates)
    }


def get_employee_requests(employee_id: str, daily_data: List[Dict] = None, start_date: date = None, end_date: date = None) -> Dict[str, Any]:
    """
    جلب طلبات الموظف من Firebase وحساب الساعات الإضافية المطلوبة وأيام الإجازة
    مع احتساب الساعات الإضافية الفعلية في الأيام المطلوبة
    يتعامل مع بنية قاعدة البيانات الجديدة: employeeId, reqDate, kind, status
    """
    try:
        # محاولة استيراد Firebase
        try:
            from firebase_config import get_db
            db = get_db()
            print(f"✅ تم الاتصال بـ Firebase بنجاح")
        except ImportError as e:
            print(f"⚠️ Firebase غير متاح: {e}")
            return {"overtime_hours": 0, "leave_days": 0, "overtime_requests": [], "leave_requests": []}
        except Exception as e:
            print(f"⚠️ خطأ في الاتصال بـ Firebase: {e}")
            return {"overtime_hours": 0, "leave_days": 0, "overtime_requests": [], "leave_requests": []}
        
        if not db:
            print("⚠️ Firebase غير متصل، سيتم تجاهل الطلبات")
            return {"overtime_hours": 0, "leave_days": 0, "overtime_requests": [], "leave_requests": []}
        
        print(f"🔍 البحث عن طلبات الموظف {employee_id} في الفترة من {start_date} إلى {end_date}")
        
        # جلب جميع الطلبات للموظف باستخدام employeeId
        requests_ref = db.collection('requests')
        
        # تجربة البحث بطرق مختلفة
        print(f"🔎 البحث باستخدام employeeId = '{employee_id}' (string)")
        query = requests_ref.where('employeeId', '==', str(employee_id))
        docs = list(query.stream())
        print(f"📋 تم العثور على {len(docs)} طلب للموظف {employee_id}")
        
        # إذا لم نجد شيء، جرب البحث بدون فلترة لرؤية البيانات
        if len(docs) == 0:
            print("🔍 لم يتم العثور على طلبات، جاري فحص جميع الطلبات...")
            all_docs = list(requests_ref.limit(10).stream())
            print(f"📋 عينة من الطلبات الموجودة ({len(all_docs)}):")
            for doc in all_docs:
                data = doc.to_dict()
                print(f"   - ID: {doc.id}, employeeId: {data.get('employeeId')}, reqDate: {data.get('reqDate')}, kind: {data.get('kind')}")
            
            # جرب البحث مرة أخرى بقيم مختلفة
            print(f"🔎 البحث مرة أخرى باستخدام employeeId = '{employee_id}'")
            query2 = requests_ref.where('employeeId', '==', employee_id)
            docs = list(query2.stream())
            print(f"📋 النتيجة الثانية: {len(docs)} طلب")
        
        overtime_hours = 0
        leave_days = 0
        overtime_dates = set()  # تواريخ الطلبات الإضافية
        leave_dates = set()     # تواريخ طلبات الإجازة
        overtime_requests = []  # تفاصيل طلبات الإضافي
        leave_requests = []     # تفاصيل طلبات الإجازة
        
        for doc in docs:
            data = doc.to_dict()
            
            # التحقق من أن الطلب نشط (غير ملغي)
            if data.get('status') != 'active':
                print(f"   ⏭️ تجاهل طلب غير نشط: {data.get('status')}")
                continue
                
            # استخدام reqDate بدلاً من date
            request_date_str = data.get('reqDate')
            if not request_date_str:
                print(f"   ⚠️ طلب بدون تاريخ: {data}")
                continue
                
            try:
                # تحويل التاريخ من string إلى date
                if isinstance(request_date_str, str):
                    request_date = datetime.strptime(request_date_str, '%Y-%m-%d').date()
                else:
                    request_date = request_date_str
                    
                # فلترة التواريخ إذا تم تحديدها (التحقق من أن الطلب في نطاق فترة المعالجة)
                # ملاحظة: إذا لم يتم تحديد فترة، سنقبل جميع الطلبات
                if start_date and end_date:
                    if request_date < start_date:
                        print(f"   ⏭️ طلب خارج النطاق (قبل): {request_date} < {start_date}")
                        continue
                    if request_date > end_date:
                        print(f"   ⏭️ طلب خارج النطاق (بعد): {request_date} > {end_date}")
                        continue
                    print(f"   ✅ طلب في النطاق: {request_date} بين {start_date} و {end_date}")
                else:
                    print(f"   ⚠️ لا توجد فترة محددة، قبول الطلب: {request_date}")
                    
                request_type = data.get('kind', '')
                request_reason = data.get('reason', '')
                
                if request_type == 'overtime':
                    overtime_dates.add(request_date)
                    overtime_requests.append({
                        'date': request_date,
                        'reason': request_reason,
                        'id': data.get('id'),
                        'supervisor': data.get('supervisor', '')
                    })
                    print(f"   ✅ طلب إضافي: {request_date} - {request_reason}")
                    
                elif request_type == 'leave':
                    leave_dates.add(request_date)
                    leave_requests.append({
                        'date': request_date,
                        'reason': request_reason,
                        'id': data.get('id'),
                        'supervisor': data.get('supervisor', '')
                    })
                    leave_days += 1
                    print(f"   ✅ طلب إجازة: {request_date} - {request_reason}")
                    
            except (ValueError, TypeError) as e:
                print(f"⚠️ خطأ في تحويل تاريخ الطلب: {request_date_str} - {e}")
                continue
        
        # حساب الساعات الإضافية الفعلية في الأيام المطلوبة
        # عندما يطلب الموظف إضافي في يوم معين، نجمع الساعات التي تزيد عن 7 ساعات في ذلك اليوم
        if daily_data and overtime_dates:
            print(f"   🧮 حساب الساعات الإضافية للأيام المطلوبة...")
            for day_record in daily_data:
                day_date = day_record.get("Date")
                if day_date in overtime_dates:
                    # حساب الساعات الإضافية في هذا اليوم (أكثر من 7 ساعات)
                    day_hours = day_record.get("DayHours", 0)
                    if day_hours > 7:
                        additional_hours = day_hours - 7
                        overtime_hours += additional_hours
                        print(f"   📅 {day_date}: {day_hours} ساعة، إضافي: {additional_hours} ساعة")
                    else:
                        print(f"   📅 {day_date}: {day_hours} ساعة، لا يوجد إضافي (أقل من 7 ساعات)")
                        
        # إذا لم توجد بيانات يومية، استخدم افتراض ساعة واحدة لكل طلب
        elif overtime_dates and not daily_data:
            overtime_hours = len(overtime_dates)  # ساعة واحدة لكل يوم إضافي مطلوب
            print(f"   ⚠️ لا توجد بيانات يومية، استخدام افتراض: {overtime_hours} ساعة")
        
        print(f"📊 موظف {employee_id}: {round(overtime_hours, 2)} ساعة إضافية مطلوبة، {leave_days} يوم إجازة مطلوب")
        
        return {
            "overtime_hours": round(overtime_hours, 2), 
            "leave_days": leave_days,
            "overtime_requests": overtime_requests,
            "leave_requests": leave_requests,
            "overtime_dates": list(overtime_dates),
            "leave_dates": list(leave_dates)
        }
        
    except Exception as e:
        print(f"❌ خطأ في جلب طلبات الموظف {employee_id}: {e}")
        import traceback
        traceback.print_exc()
        return {"overtime_hours": 0, "leave_days": 0, "overtime_requests": [], "leave_requests": []}


def analyze_file(path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """
    تحليل ملف الحضور وإرجاع معلومات أساسية عنه
    """
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[sheet_name] if sheet_name else wb.worksheets[0]
        
        rows = list(ws.iter_rows(values_only=False))
        nrows = len(rows)
        
        # البحث عن الموظفين
        employees_found = 0
        file_format = "unknown"
        first_date = None
        last_date = None
        all_dates = set()
        
        r = 0
        while r < nrows:
            row_cells = list(rows[r])
            header = parse_employee_header(row_cells)
            if not header:
                r += 1
                continue
                
            employees_found += 1
            
            # تحديد نوع الملف
            if r + 1 < nrows:
                if detect_is_timecard_header(rows[r+1]):
                    file_format = "timecard"
                else:
                    file_format = "legacy"
            
            # جمع التواريخ من بيانات الموظف
            data_start = r + 1
            if file_format == "timecard":
                data_start = r + 2  # تخطي header الـ timecard
            
            # البحث عن التواريخ في البيانات
            for i in range(data_start, min(data_start + 50, nrows)):  # فحص أول 50 صف فقط
                if i >= nrows:
                    break
                    
                row_cells = list(rows[i])
                if not row_cells or not row_cells[0].value:
                    break
                    
                # محاولة استخراج التاريخ من العمود الأول
                try:
                    cell_value = row_cells[0].value
                    if isinstance(cell_value, datetime):
                        date_obj = cell_value.date()
                        all_dates.add(date_obj)
                    elif isinstance(cell_value, str):
                        # محاولة تحويل النص إلى تاريخ
                        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y']:
                            try:
                                date_obj = datetime.strptime(cell_value, fmt).date()
                                all_dates.add(date_obj)
                                break
                            except ValueError:
                                continue
                except:
                    continue
            
            # الانتقال للموظف التالي - البحث عن الموظف التالي
            r = data_start
            # البحث عن نهاية بيانات الموظف الحالي
            while r < nrows:
                row_cells = list(rows[r])
                if not row_cells or not row_cells[0].value:
                    r += 1
                    break
                # إذا وجدنا موظف جديد، توقف
                if parse_employee_header(row_cells):
                    break
                r += 1
        
        # حساب أول وآخر تاريخ
        if all_dates:
            first_date = min(all_dates)
            last_date = max(all_dates)
            period_days = (last_date - first_date).days + 1
        else:
            period_days = 0
        
        # تحديد نوع الملف بشكل أفضل
        if file_format == "unknown" and employees_found > 0:
            file_format = "legacy"  # افتراض legacy إذا لم نتمكن من التحديد
        
        # جلب إحصائيات الطلبات للفترة المحددة
        overtime_requests_count = 0
        leave_requests_count = 0
        
        if first_date and last_date:
            try:
                # جلب جميع الطلبات النشطة للفترة
                all_requests = get_all_active_requests(first_date, last_date)
                
                # حساب إجمالي الطلبات
                for employee_id, requests_data in all_requests.items():
                    overtime_requests_count += len(requests_data.get('overtime_requests', []))
                    leave_requests_count += len(requests_data.get('leave_requests', []))
                
                print(f"📊 إحصائيات الطلبات للفترة {first_date} - {last_date}:")
                print(f"   - طلبات الإضافي: {overtime_requests_count}")
                print(f"   - طلبات الإجازة: {leave_requests_count}")
                
            except Exception as e:
                print(f"⚠️ خطأ في جلب إحصائيات الطلبات: {e}")
        
        return {
            "employees_count": employees_found,
            "file_format": file_format,
            "first_date": first_date.strftime('%Y-%m-%d') if first_date else None,
            "last_date": last_date.strftime('%Y-%m-%d') if last_date else None,
            "period_days": period_days,
            "total_rows": nrows,
            "sheet_name": ws.title,
            "dates_found": len(all_dates),
            "overtime_requests_count": overtime_requests_count,
            "leave_requests_count": leave_requests_count,
            "success": True
        }
        
    except Exception as e:
        print(f"❌ خطأ في تحليل الملف: {e}")
        import traceback
        traceback.print_exc()
        
        return {
            "success": False,
            "error": str(e),
            "employees_count": 0,
            "file_format": "unknown",
            "first_date": None,
            "last_date": None,
            "period_days": 0,
            "total_rows": 0,
            "sheet_name": "غير معروف",
            "dates_found": 0
        }


def parse_args():
    p = argparse.ArgumentParser(description="Process attendance Excel and compute per-employee metrics.")
    p.add_argument("input", help="Path to the input Excel file (as exported from attendance system)")
    p.add_argument("--sheet", help="Worksheet name to read (default: first sheet)", default=None)
    p.add_argument("--target-days", type=int, required=True, help="Target number of workdays in the period (e.g., 26)")
    p.add_argument(
        "--holidays",
        type=str,
        default="",
        help="Comma-separated list of official holiday dates in YYYY-MM-DD (e.g., 2025-09-05,2025-09-10)",
    )
    p.add_argument(
        "--special-days",
        type=str,
        default="",
        help="Comma-separated list of exceptional dates (YYYY-MM-DD) where absence should NOT count as AbsentDays",
    )
    p.add_argument("--output", default="attendance_summary.xlsx", help="Output Excel file path for the summary")
    p.add_argument("--output-daily", default="attendance_daily.xlsx", help="Output Excel file path for per-day details")
    p.add_argument("--out-dir", default="", help="If provided, both outputs will be written into this folder as Summary_YYYYMMDDHHMMSS.xlsx and Daily_YYYYMMDDHHMMSS.xlsx")
    p.add_argument("--cutoff-hour", type=int, default=7, help="Overnight cutoff hour. Times before this hour at the start of a day are treated as previous day's last punch (default: 7)")
    p.add_argument("--format", choices=["auto", "legacy", "timecard"], default="auto", help="Input format: auto-detect, legacy (Date|First Punch|Last Punch), or timecard (Date|Times|Time list)")
    p.add_argument("--dup-threshold-minutes", type=int, default=60, help="When two consecutive punches in the same day are closer than this number of minutes, drop the newer as a duplicate (default: 60)")
    p.add_argument("--assume-missing-exit-hours", type=float, default=5.0, help="If a day ends with an unmatched entry (no exit), assume this many hours for the missing exit (default: 5.0)")
    p.add_argument(
        "--overtime-positive-only",
        action="store_true",
        help="If set, negative overtime will be clipped to 0 (default behavior).",
    )
    p.add_argument(
        "--allow-negative-overtime",
        action="store_true",
        help="If set, overtime may be negative. Overrides --overtime-positive-only.",
    )
    return p.parse_args()


def parse_holidays(hol_str: str) -> set:
    holidays = set()
    if not hol_str:
        return holidays
    for part in hol_str.split(','):
        s = part.strip()
        if not s:
            continue
        try:
            holidays.add(datetime.strptime(s, "%Y-%m-%d").date())
        except ValueError:
            raise ValueError(f"Invalid holiday date format: {s}. Expected YYYY-MM-DD")
    return holidays


def cell_text(cell: Optional[Cell]) -> str:
    if cell is None:
        return ""
    v = cell.value
    return "" if v is None else str(v).strip()


def parse_employee_header(row_cells: List[Cell]) -> Optional[Dict[str, str]]:
    # Support two formats:
    # 1) All-in-one A cell: 'Employee ID: X, First Name: Y, Department: Z'
    # 2) Split across A/B/C cells respectively
    a = cell_text(row_cells[0]) if len(row_cells) > 0 else ""
    
    # البحث عن Employee ID بطرق مختلفة
    a_lower = a.lower()
    if not (a.startswith(EMPLOYEE_MARKER) or 
            "employee id:" in a_lower or 
            "employee id :" in a_lower or
            "employeeid:" in a_lower or
            a_lower.startswith("employee")):
        return None
    header = {"EmployeeID": "", "Name": None, "Department": None}
    # First, try to parse from A if it contains comma-separated key:value pairs
    try:
        tokens = [t.strip() for t in a.split(',') if t.strip()]
        for tok in tokens:
            low = tok.lower()
            if low.startswith("employee id") and ":" in tok:
                header["EmployeeID"] = tok.split(":", 1)[1].strip()
            elif low.startswith("first name") and ":" in tok:
                header["Name"] = tok.split(":", 1)[1].strip()
            elif low.startswith("department") and ":" in tok:
                header["Department"] = tok.split(":", 1)[1].strip()
    except Exception:
        pass
    # If some fields are still missing, also check B and C cells (some exports split them)
    if header.get("Name") is None and len(row_cells) > 1:
        b = cell_text(row_cells[1])
        if b.lower().startswith("first name") and ":" in b:
            header["Name"] = b.split(":", 1)[1].strip()
    if header.get("Department") is None and len(row_cells) > 2:
        c = cell_text(row_cells[2])
        if c.lower().startswith("department") and ":" in c:
            header["Department"] = c.split(":", 1)[1].strip()
    # As a last resort, ensure EmployeeID is filled from A
    if not header.get("EmployeeID"):
        header["EmployeeID"] = a.split(":", 1)[1].strip() if ":" in a else a
    return header


def to_date(val) -> Optional[date]:
    if val is None or str(val).strip() == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    # Try common formats
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # As last resort, try Excel serial? Not handled here.
    return None


def to_time(val) -> Optional[time]:
    if val is None or str(val).strip() == "":
        return None
    if isinstance(val, datetime):
        return val.time()
    if isinstance(val, time):
        return val
    # Some exports store as "HH:MM" or "H:MM"
    s = str(val).strip()
    # Accept 24-hour times
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    # If numeric (Excel time as fraction of day)
    try:
        f = float(s)
        # Excel time: 1.0 == 24 hours
        seconds = int(round(f * 24 * 3600))
        seconds %= 24 * 3600
        return (datetime.min + timedelta(seconds=seconds)).time()
    except Exception:
        return None


def hours_decimal_from_times(first: Optional[time], last: Optional[time]) -> float:
    if not first or not last:
        return 0.0
    dt0 = datetime.combine(date(2000,1,1), first)
    dt1 = datetime.combine(date(2000,1,1), last)
    delta = dt1 - dt0
    # If negative (overnight not expected here), treat as 0
    if delta.total_seconds() < 0:
        return 0.0
    return round(delta.total_seconds() / 3600.0, 4)


def hours_decimal_between(d0: date, t0: time, d1: date, t1: time) -> float:
    dt0 = datetime.combine(d0, t0)
    dt1 = datetime.combine(d1, t1)
    if dt1 < dt0:
        # crosses midnight
        dt1 = dt1 + timedelta(days=1)
    return round((dt1 - dt0).total_seconds() / 3600.0, 4)


def detect_is_timecard_header(cells: List[Cell]) -> bool:
    a = cell_text(cells[0]) if len(cells) > 0 else ""
    b = cell_text(cells[1]) if len(cells) > 1 else ""
    c = cell_text(cells[2]) if len(cells) > 2 else ""
    return a.lower() == "date" and b.lower().startswith("times") and c.lower() == "time"


def process_timecard_section(rows: List[List[Cell]], start_idx: int, holidays: set, cutoff_hour: int, dup_threshold_minutes: int, assume_missing_exit_hours: float) -> (int, Dict[str, Any]):
    """Process a single employee block in timecard format starting at start_idx where row[start] is employee header.
    Returns (next_index, result_dict).
    """
    header = parse_employee_header(rows[start_idx])
    print(f"Processing timecard section for employee: {header}")
    i = start_idx + 1
    n = len(rows)
    # Expect header line Date | Times | Time
    if i < n and detect_is_timecard_header(rows[i]):
        i += 1
    # Collect date->list[time]
    day_map: Dict[date, List[time]] = {}
    while i < n:
        a_text = cell_text(rows[i][0]) if rows[i] else ""
        if a_text.startswith(EMPLOYEE_MARKER):
            break
        d = to_date(rows[i][0].value) if len(rows[i]) > 0 else None
        if not d:
            i += 1
            continue
        # times list is column 3 (index 2)
        time_cell_val = rows[i][2].value if len(rows[i]) > 2 else None
        tlist: List[time] = []
        if time_cell_val is not None:
            s = str(time_cell_val).strip()
            if s:
                for tok in s.split(','):
                    tt = to_time(tok.strip())
                    if tt:
                        tlist.append(tt)
        if tlist:
            tlist.sort()
            day_map.setdefault(d, []).extend(tlist)
        i += 1

    # Apply cutoff shift: early punches before cutoff at start of a day move to previous day as last punches
    dates_sorted = sorted(day_map.keys())
    prev_date: Optional[date] = None
    for d in dates_sorted:
        times_today = day_map.get(d, [])
        # Collapse near-duplicate early punches before moving (keep the earliest, drop subsequent within threshold)
        while (
            len(times_today) >= 2
            and times_today[0].hour < cutoff_hour
            and times_today[1].hour < cutoff_hour
        ):
            dt0 = datetime.combine(d, times_today[0])
            dt1 = datetime.combine(d, times_today[1])
            delta_min = abs((dt1 - dt0).total_seconds()) / 60.0
            if delta_min < dup_threshold_minutes:
                # drop the newer (second)
                times_today.pop(1)
            else:
                break
        if times_today and times_today[0].hour < cutoff_hour:
            early = times_today.pop(0)
            if prev_date is None:
                # First day in the processed range: treat as exit from previous period -> drop it
                pass
            else:
                day_map.setdefault(prev_date, []).append(early)
        if not times_today:
            # if emptied, remove
            if d in day_map and len(day_map[d]) == 0:
                del day_map[d]
        prev_date = d

    # Recompute sorted dates after possible removals/additions
    dates_sorted = sorted(day_map.keys())

    work_days = 0
    total_hours = 0.0
    overtime_sum = 0.0
    delay_sum = 0.0
    worked_dates: List[date] = []
    daily: List[Dict[str, Any]] = []
    for d in dates_sorted:
        # Preserve insertion order to keep cross-midnight moved punches at the end of previous day
        tl = list(day_map[d])
        # Remove near-duplicate punches (keep older, drop newer if within threshold minutes)
        if tl:
            filtered = []
            prev_dt = None
            for t in tl:
                cur_dt = datetime.combine(d, t)
                if prev_dt is None:
                    filtered.append(t)
                    prev_dt = cur_dt
                else:
                    delta = (cur_dt - prev_dt).total_seconds()
                    if delta < 0:
                        # if out-of-order for any reason, use absolute delta
                        delta = abs(delta)
                    if delta >= dup_threshold_minutes * 60:
                        filtered.append(t)
                        prev_dt = cur_dt
            tl = filtered
        # pair sequentially
        paired = 0
        j = 0
        day_hours = 0.0
        while j + 1 < len(tl):
            h = hours_decimal_between(d, tl[j], d, tl[j+1])
            total_hours += h
            day_hours += h
            paired += 1
            j += 2
        assumed_exit = 0
        # If there's an unmatched last entry, assume fixed hours only if that unmatched punch is an entry (>= cutoff)
        if j < len(tl):
            last_punch = tl[-1]
            if last_punch.hour >= cutoff_hour:
                day_hours += assume_missing_exit_hours
                total_hours += assume_missing_exit_hours
                paired += 1
                assumed_exit = 1
        if paired > 0:
            work_days += 1
            worked_dates.append(d)
            # per-day over/under vs 7 hours baseline
            overtime_sum += max(0.0, day_hours - 7.0)
            delay_sum += max(0.0, 7.0 - day_hours)
        daily.append({
            "Date": d,
            "TimesCount": len(tl),
            "TimesList": ",".join(t.strftime("%H:%M:%S") for t in tl),
            "DayHours": round(day_hours, 4),
            "Worked": 1 if paired > 0 else 0,
            "IsHoliday": 1 if d in holidays else 0,
            "ShiftsCount": paired,
            "AssumedExit": assumed_exit,
            "DayOvertimeHours": round(max(0.0, day_hours - 7.0), 4),
            "DayDelayHours": round(max(0.0, 7.0 - day_hours), 4),
        })

    worked_on_holidays = sum(1 for dd in worked_dates if dd in holidays)
    result = {
        "EmployeeID": header.get("EmployeeID"),
        "Name": header.get("Name"),
        "Department": header.get("Department"),
        # WorkDays etc computed at higher level where target_days is known
        "_work_days": work_days,
        "_total_hours": round(total_hours, 4),
        "_worked_dates": worked_dates,
        "_daily": daily,
        "_overtime_sum": round(overtime_sum, 4),
        "_delay_sum": round(delay_sum, 4),
    }
    return i, result


def process_legacy_section(rows: List[List[Cell]], start_idx: int, holidays: set, assume_missing_exit_hours: float) -> (int, Dict[str, Any]):
    header = parse_employee_header(rows[start_idx])
    print(f"Processing legacy section for employee: {header}")
    i = start_idx + 1
    n = len(rows)
    work_days = 0
    total_hours = 0.0
    worked_dates: List[date] = []
    daily: List[Dict[str, Any]] = []
    overtime_sum = 0.0
    delay_sum = 0.0
    while i < n:
        a_text = cell_text(rows[i][0]) if rows[i] else ""
        if a_text.startswith(EMPLOYEE_MARKER):
            break
        # Skip header line
        if a_text.lower() == "date":
            i += 1
            continue
        d = to_date(rows[i][0].value) if len(rows[i]) > 0 else None
        if not d:
            i += 1
            continue
        first_punch = to_time(rows[i][2].value) if len(rows[i]) > 2 else None
        last_punch = to_time(rows[i][3].value) if len(rows[i]) > 3 else None
        if first_punch and last_punch:
            work_days += 1
            worked_dates.append(d)
            dh = hours_decimal_between(d, first_punch, d, last_punch)
            total_hours += dh
            overtime_sum += max(0.0, dh - 7.0)
            delay_sum += max(0.0, 7.0 - dh)
            daily.append({
                "Date": d,
                "TimesCount": 2,
                "TimesList": ",".join([first_punch.strftime("%H:%M:%S"), last_punch.strftime("%H:%M:%S")]),
                "DayHours": round(dh, 4),
                "Worked": 1,
                "IsHoliday": 1 if d in holidays else 0,
                "ShiftsCount": 1,
                "AssumedExit": 0,
                "DayOvertimeHours": round(max(0.0, dh - 7.0), 4),
                "DayDelayHours": round(max(0.0, 7.0 - dh), 4),
            })
        else:
            # if only entry exists (first punch), assume fixed hours
            if first_punch and not last_punch:
                work_days += 1
                worked_dates.append(d)
                dh = assume_missing_exit_hours
                total_hours += dh
                overtime_sum += max(0.0, dh - 7.0)
                delay_sum += max(0.0, 7.0 - dh)
                daily.append({
                    "Date": d,
                    "TimesCount": 1,
                    "TimesList": first_punch.strftime("%H:%M:%S"),
                    "DayHours": round(dh, 4),
                    "Worked": 1,
                    "IsHoliday": 1 if d in holidays else 0,
                    "ShiftsCount": 1,
                    "AssumedExit": 1,
                    "DayOvertimeHours": round(max(0.0, dh - 7.0), 4),
                    "DayDelayHours": round(max(0.0, 7.0 - dh), 4),
                })
            else:
                # add a daily row with zero hours if there's a date but incomplete or no punches
                daily.append({
                    "Date": d,
                    "TimesCount": (1 if (first_punch or last_punch) else 0),
                    "TimesList": ",".join([t.strftime("%H:%M:%S") for t in [first_punch, last_punch] if t]),
                    "DayHours": 0.0,
                    "Worked": 0,
                    "IsHoliday": 1 if d in holidays else 0,
                    "ShiftsCount": 0,
                    "AssumedExit": 0,
                    "DayOvertimeHours": 0.0,
                    "DayDelayHours": 0.0,
                })
        i += 1
    result = {
        "EmployeeID": header.get("EmployeeID"),
        "Name": header.get("Name"),
        "Department": header.get("Department"),
        "_work_days": work_days,
        "_total_hours": round(total_hours, 4),
        "_worked_dates": worked_dates,
        "_daily": daily,
    }
    return i, result


def process_workbook(path: str, sheet_name: Optional[str], target_days: int, holidays: set, special_days: set = None, fmt: str = "auto", cutoff_hour: int = 7, dup_threshold_minutes: int = 60, assume_missing_exit_hours: float = 5.0) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    print(f"Starting file processing: {path}")
    print(f"Processing parameters: target_days={target_days}, fmt={fmt}, cutoff_hour={cutoff_hour}")
    
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]
    
    print(f"Sheet info: {ws.title}, rows: {ws.max_row}, columns: {ws.max_column}")

    rows = list(ws.iter_rows(values_only=False))
    r = 0
    nrows = len(rows)
    results: List[Dict[str, Any]] = []
    daily_rows: List[Dict[str, Any]] = []
    
    print(f"Searching for employees in {nrows} rows...")
    employees_found = 0
    
    # 🚀 جلب جميع الطلبات النشطة مرة واحدة لتحسين الأداء
    print(f"🚀 جلب جميع الطلبات النشطة من قاعدة البيانات...")
    
    # محاولة تحديد فترة التواريخ من الملف لتحسين الفلترة
    # سنجلب جميع الطلبات أولاً ثم نفلترها لاحقاً حسب فترة كل موظف
    all_requests = get_all_active_requests()
    print(f"✅ تم جلب الطلبات لـ {len(all_requests)} موظف")

    while r < nrows:
        row_cells = list(rows[r])
        header = parse_employee_header(row_cells)
        if not header:
            # Print row content for diagnosis
            if r < 20:  # Only first 20 rows to avoid spam
                cell_content = cell_text(row_cells[0]) if len(row_cells) > 0 else ""
                if cell_content:
                    print(f"  Row {r+1}: '{cell_content}' - No Employee ID")
            r += 1
            continue
            
        employees_found += 1
        print(f"Found employee #{employees_found} in row {r+1}: {header}")
        
        # Decide which section parser to use
        next_idx = r
        if fmt == "timecard" or (fmt == "auto" and r + 1 < nrows and detect_is_timecard_header(rows[r+1])):
            next_idx, partial = process_timecard_section(rows, r, holidays, cutoff_hour, dup_threshold_minutes, assume_missing_exit_hours)
        else:
            next_idx, partial = process_legacy_section(rows, r, holidays, assume_missing_exit_hours)

        work_days = partial.get("_work_days", 0)
        total_hours = partial.get("_total_hours", 0.0)
        worked_dates = partial.get("_worked_dates", [])
        worked_on_holidays = sum(1 for d in worked_dates if d in holidays)
        non_holiday_work_days = max(0, work_days - worked_on_holidays)
        # Special days: if absent on these dates, do NOT count as AbsentDays
        sdays = special_days or set()
        special_absent_ignored = sum(1 for sd in sdays if sd not in worked_dates)
        # AbsentDays (all): difference vs target minus ignored special-absent days
        absent_all = max(0, target_days - work_days - special_absent_ignored)
        # AbsentDaysExclHolidays: excludes holidays entirely
        absent_excl_holidays = max(0, target_days - non_holiday_work_days)
        # Extra days per requested rule: days beyond target (regardless of holiday) + all worked holidays
        extra_days = max(0, work_days - target_days) + worked_on_holidays
        overtime = partial.get("_overtime_sum", max(0.0, total_hours - 7.0 * work_days))
        delay_hours = partial.get("_delay_sum", max(0.0, 7.0 * work_days - total_hours))
        # Count days where exit was assumed (5 hours by default)
        assumed_exit_days = sum(1 for d in partial.get("_daily", []) if d.get("AssumedExit") == 1)
        
        # جلب طلبات الموظف من البيانات المُحملة مسبقاً (أكثر كفاءة)
        employee_id = partial.get("EmployeeID")
        daily_data = partial.get("_daily", [])
        
        # استخدام الدالة الجديدة الأكثر كفاءة
        requests_data = process_employee_requests_from_cache(employee_id, daily_data, all_requests)
        
        res_row = {
            "EmployeeID": employee_id,
            "Name": partial.get("Name"),
            "Department": partial.get("Department"),
            "WorkDays": work_days,
            # AbsentDays: كل الغياب مقارنة بالـ TargetDays (يشمل العطل لو لم يعمل فيها)
            "AbsentDays": absent_all,
            "AbsentDaysExclHolidays": absent_excl_holidays,
            "ExtraDays": extra_days,
            "TotalHours": round(total_hours, 4),
            "OvertimeHours": round(overtime, 4),
            "DelayHours": round(delay_hours, 4),
            "WorkedOnHolidays": worked_on_holidays,
            "AssumedExitDays": assumed_exit_days,
            # معلومات الطلبات المعتمدة من قاعدة البيانات
            "RequestedOvertimeHours": requests_data.get("overtime_hours", 0),
            "RequestedLeaveDays": requests_data.get("leave_days", 0),
            "OvertimeRequestsCount": len(requests_data.get("overtime_requests", [])),
            "LeaveRequestsCount": len(requests_data.get("leave_requests", [])),
            # تفاصيل إضافية للطلبات (للتصدير المفصل)
            "OvertimeRequestsDates": "; ".join([str(req['date']) for req in requests_data.get("overtime_requests", [])]),
            "LeaveRequestsDates": "; ".join([str(req['date']) for req in requests_data.get("leave_requests", [])]),
            "OvertimeRequestsReasons": "; ".join([req['reason'] for req in requests_data.get("overtime_requests", [])]),
            "LeaveRequestsReasons": "; ".join([req['reason'] for req in requests_data.get("leave_requests", [])]),
        }
        results.append(res_row)
        # Attach employee info to daily rows and collect
        for drow in partial.get("_daily", []):
            day_date = drow.get("Date")
            
            # التحقق من وجود طلبات في هذا اليوم
            has_overtime_request = day_date in requests_data.get("overtime_dates", [])
            has_leave_request = day_date in requests_data.get("leave_dates", [])
            
            # البحث عن تفاصيل الطلب في هذا اليوم
            overtime_reason = ""
            leave_reason = ""
            
            for req in requests_data.get("overtime_requests", []):
                if req['date'] == day_date:
                    overtime_reason = req['reason']
                    break
                    
            for req in requests_data.get("leave_requests", []):
                if req['date'] == day_date:
                    leave_reason = req['reason']
                    break
            
            row = {
                "EmployeeID": res_row["EmployeeID"],
                "Name": res_row["Name"],
                "Department": res_row["Department"],
                "Date": drow.get("Date"),
                "TimesCount": drow.get("TimesCount"),
                "TimesList": drow.get("TimesList"),
                "DayHours": drow.get("DayHours"),
                "IsHoliday": drow.get("IsHoliday"),
                "DayOvertimeHours": drow.get("DayOvertimeHours"),
                "DayDelayHours": drow.get("DayDelayHours"),
                # معلومات الطلبات لهذا اليوم
                "HasOvertimeRequest": has_overtime_request,
                "HasLeaveRequest": has_leave_request,
                "OvertimeRequestReason": overtime_reason,
                "LeaveRequestReason": leave_reason,
            }
            daily_rows.append(row)
        r = next_idx
    
    print(f"Processing completed:")
    print(f"   - Employees found: {employees_found}")
    print(f"   - Summary results: {len(results)}")
    print(f"   - Daily records: {len(daily_rows)}")
    
    # إحصائيات الطلبات
    total_overtime_requests = sum(res.get("OvertimeRequestsCount", 0) for res in results)
    total_leave_requests = sum(res.get("LeaveRequestsCount", 0) for res in results)
    total_requested_overtime_hours = sum(res.get("RequestedOvertimeHours", 0) for res in results)
    total_requested_leave_days = sum(res.get("RequestedLeaveDays", 0) for res in results)
    
    print(f"📊 إحصائيات الطلبات:")
    print(f"   - طلبات إضافي: {total_overtime_requests} طلب")
    print(f"   - طلبات إجازة: {total_leave_requests} طلب")
    print(f"   - ساعات إضافية مطلوبة: {total_requested_overtime_hours} ساعة")
    print(f"   - أيام إجازة مطلوبة: {total_requested_leave_days} يوم")
    
    if results:
        print(f"   - First result: {results[0]}")
    if daily_rows:
        print(f"   - First daily record: {daily_rows[0]}")
    
    return results, daily_rows


def write_summary(output_path: str, results: List[Dict[str, Any]], config: Dict[str, Any]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    headers = [
        "Employee ID",
        "Employee Name", 
        "Department",
        "Work Days",
        "Absent Days",
        "Worked on Holidays",
        "Extra Days",
        "Total Hours",
        "Overtime Hours",
        "Requested Overtime Hours",
        "Delay Hours",
        "Overtime Requests Count",
        "Leave Requests Count",
        "Missing Punches",
    ]
    ws.append(headers)
    for row in results:
        ws.append([
            row.get("EmployeeID"),                      # Employee ID
            row.get("Name"),                            # Employee Name
            row.get("Department"),                      # Department
            row.get("WorkDays"),                        # Work Days
            row.get("AbsentDays"),                      # Absent Days
            row.get("WorkedOnHolidays"),                # Worked on Holidays
            row.get("ExtraDays"),                       # Extra Days
            row.get("TotalHours"),                      # Total Hours
            row.get("OvertimeHours"),                   # Overtime Hours
            row.get("RequestedOvertimeHours", 0),       # Requested Overtime Hours
            row.get("DelayHours"),                      # Delay Hours
            row.get("OvertimeRequestsCount", 0),        # Overtime Requests Count
            row.get("LeaveRequestsCount", 0),           # Leave Requests Count
            row.get("AssumedExitDays"),                 # Missing Punches
        ])
    # Add a config sheet
    ws2 = wb.create_sheet("Config")
    ws2.append(["TargetDays", config.get("target_days")])
    ws2.append(["Holidays", ",".join(sorted(d.strftime('%Y-%m-%d') for d in config.get("holidays", set())))])
    if config.get("special_days") is not None:
        ws2.append(["SpecialDays", ",".join(sorted(d.strftime('%Y-%m-%d') for d in config.get("special_days", set())))])
    wb.save(output_path)


def write_daily_details(output_path: str, daily_rows: List[Dict[str, Any]]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily"
    headers = [
        "EmployeeID",
        "Name",
        "Department",
        "Date",
        "TimesCount",
        "TimesList",
        "DayHours",
        "IsHoliday",
        "DayOvertimeHours",
        "DayDelayHours",
        "Has Overtime Request",
        "Has Leave Request",
        "Overtime Request Reason",
        "Leave Request Reason",
    ]
    ws.append(headers)
    # Sort by employee then date
    daily_rows_sorted = sorted(daily_rows, key=lambda x: (str(x.get("EmployeeID")), x.get("Date") or datetime(1900,1,1).date()))
    for row in daily_rows_sorted:
        ws.append([
            row.get("EmployeeID"),
            row.get("Name"),
            row.get("Department"),
            row.get("Date").strftime('%Y-%m-%d') if row.get("Date") else None,
            row.get("TimesCount"),
            row.get("TimesList"),
            row.get("DayHours"),
            row.get("IsHoliday"),
            row.get("DayOvertimeHours"),
            row.get("DayDelayHours"),
            row.get("HasOvertimeRequest", False),
            row.get("HasLeaveRequest", False),
            row.get("OvertimeRequestReason", ""),
            row.get("LeaveRequestReason", ""),
        ])
    wb.save(output_path)


def main():
    args = parse_args()
    # Decide overtime clipping
    clip_positive = True
    if args.allow_negative_overtime:
        clip_positive = False
    holidays = parse_holidays(args.holidays)
    special_days = parse_holidays(args.special_days) if getattr(args, 'special_days', "") else set()
    results, daily_rows = process_workbook(args.input, args.sheet, args.target_days, holidays, special_days, fmt=args.format, cutoff_hour=args.cutoff_hour, dup_threshold_minutes=args.dup_threshold_minutes, assume_missing_exit_hours=args.assume_missing_exit_hours)
    if clip_positive:
        for r in results:
            if r["OvertimeHours"] < 0:
                r["OvertimeHours"] = 0.0
    # Determine output paths
    out_summary = args.output
    out_daily = args.output_daily
    if args.out_dir:
        ts = datetime.now().strftime('%Y%m%d%H%M%S')
        out_dir = args.out_dir
        try:
            os.makedirs(out_dir, exist_ok=True)
        except Exception:
            pass
        out_summary = os.path.join(out_dir, f"Summary_{ts}.xlsx")
        out_daily = os.path.join(out_dir, f"Daily_{ts}.xlsx")
    write_summary(out_summary, results, {"target_days": args.target_days, "holidays": holidays, "special_days": special_days})
    write_daily_details(out_daily, daily_rows)
    print(f"Processed {len(results)} employees. Summary: {out_summary} | Daily: {out_daily}")


if __name__ == "__main__":
    main()
