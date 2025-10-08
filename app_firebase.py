import io
import os
import sys
import tempfile
import zipfile
from datetime import datetime, timedelta
from functools import wraps
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
# Optional deps: Flask-Compress and Flask-Caching
try:
    from flask_compress import Compress  # type: ignore
except Exception:
    class Compress:  # no-op fallback
        def __init__(self, *args, **kwargs):
            pass
        def init_app(self, *args, **kwargs):
            pass
try:
    from flask_caching import Cache  # type: ignore
except Exception:
    # Minimal no-op cache with get/set interface
    class _NoopCache:
        def __init__(self, *args, **kwargs):
            pass
        def get(self, key):
            return None
        def set(self, key, value, timeout=None):
            return True
    def Cache(*args, **kwargs):
        return _NoopCache()
from werkzeug.security import check_password_hash, generate_password_hash
try:
    from google.api_core.exceptions import ResourceExhausted
except Exception:  # library not available at type-check time
    class ResourceExhausted(Exception):
        pass
import jwt
from firebase_config import (
    initialize_firebase,
    get_user_by_username,
    create_user,
    get_all_users,
    get_pending_users,
    add_pending_user,
    approve_pending_user,
    reject_pending_user,
    delete_user,
    create_request,
    get_latest_requests,
    cancel_request
)

# Ensure we can import attendance_processor from parent directory
CUR_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.abspath(os.path.join(CUR_DIR, os.pardir))
if PARENT_DIR not in sys.path:
    sys.path.insert(0, PARENT_DIR)

from attendance_processor import (
    process_workbook,
    parse_holidays,
)
from openpyxl import Workbook

# قاموس الترجمات
TRANSLATIONS = {
    'ar': {
        'summary_title': 'ملخص الحضور',
        'daily_title': 'التفاصيل اليومية',
        'times_title': 'جميع الأوقات',
        'summary_filename': 'تقرير_الملخص.xlsx',
        'daily_filename': 'التفاصيل_اليومية.xlsx',
        'times_filename': 'جميع_الأوقات.xlsx',
        'zip_filename': 'تقارير_الحضور',
        'summary_headers': [
                'رقم الموظف', 'اسم الموظف', 'القسم', 'أيام الحضور', 'أيام الغياب',
                'عمل في العطل', 'أيام إضافية', 'ساعات العمل', 'ساعات الإضافي',
                'ساعات إضافي مطلوبة', 'ساعات التأخير', 'عدد طلبات الإضافي', 'عدد طلبات الإجازة', 'البصمات المنسية'
            ],
        'daily_headers': [
            'رقم الموظف', 'اسم الموظف', 'القسم', 'التاريخ', 'أول دخول', 'آخر خروج',
            'ساعات العمل', 'ساعات الإضافي', 'ساعات التأخير', 'عدد مرات الدخول/الخروج', 'يوم عطلة',
            'يوجد طلب إضافي', 'يوجد طلب إجازة', 'سبب طلب الإضافي', 'سبب طلب الإجازة'
        ],
        'times_headers': [
            'رقم الموظف', 'اسم الموظف', 'القسم', 'التاريخ', 'جميع أوقات الدخول والخروج', 'عدد المرات', 'يوم عطلة'
        ],
        'yes': 'نعم',
        'no': 'لا',
        'no_data': 'لا توجد بيانات',
        'check_format': 'تحقق من تنسيق الملف',
        'no_daily_data': 'لا توجد بيانات يومية'
    },
    'en': {
        'summary_title': 'Attendance Summary',
        'daily_title': 'Daily Details',
        'times_title': 'All Times',
        'summary_filename': 'Summary_Report.xlsx',
        'daily_filename': 'Daily_Details.xlsx',
        'times_filename': 'All_Times.xlsx',
        'zip_filename': 'attendance_reports',
        'summary_headers': [
            'Employee ID', 'Employee Name', 'Department', 'Work Days', 'Absent Days',
            'Worked on Holidays', 'Extra Days', 'Total Hours', 'Overtime Hours',
            'Requested Overtime Hours', 'Delay Hours', 'Overtime Requests Count', 'Leave Requests Count', 'Missing Punches'
        ],
        'daily_headers': [
            'Employee ID', 'Employee Name', 'Department', 'Date', 'First In', 'Last Out',
            'Work Hours', 'Overtime Hours', 'Delay Hours', 'Times Count', 'Holiday',
            'Has Overtime Request', 'Has Leave Request', 'Overtime Request Reason', 'Leave Request Reason'
        ],
        'times_headers': [
            'Employee ID', 'Employee Name', 'Department', 'Date', 'All Times', 'Times Count', 'Holiday'
        ],
        'yes': 'Yes',
        'no': 'No',
        'no_data': 'No data',
        'check_format': 'Check file format',
        'no_daily_data': 'No daily data'
    }
}

def get_translation(language, key):
    """الحصول على الترجمة المناسبة"""
    return TRANSLATIONS.get(language, TRANSLATIONS['ar']).get(key, key)

def get_employee_overtime_requests(employee_id, start_date, end_date):
    """جلب طلبات الإضافي المعتمدة للموظف في فترة معينة"""
    try:
        from firebase_config import db
        if not db:
            return 0.0
        
        # البحث عن طلبات الإضافي المعتمدة
        requests_ref = db.collection('requests')
        query = requests_ref.where('employeeId', '==', str(employee_id)) \
                           .where('type', '==', 'overtime') \
                           .where('status', '==', 'approved') \
                           .where('date', '>=', start_date) \
                           .where('date', '<=', end_date)
        
        total_hours = 0.0
        for doc in query.stream():
            data = doc.to_dict()
            hours = float(data.get('hours', 0))
            total_hours += hours
        
        return total_hours
    except Exception as e:
        print(f"خطأ في جلب طلبات الإضافي للموظف {employee_id}: {e}")
        return 0.0

def get_employee_leave_requests(employee_id, start_date, end_date):
    """جلب طلبات الإجازة المعتمدة للموظف في فترة معينة"""
    try:
        from firebase_config import db
        if not db:
            return 0
        
        # البحث عن طلبات الإجازة المعتمدة
        requests_ref = db.collection('requests')
        query = requests_ref.where('employeeId', '==', str(employee_id)) \
                           .where('type', '==', 'leave') \
                           .where('status', '==', 'approved') \
                           .where('startDate', '>=', start_date) \
                           .where('endDate', '<=', end_date)
        
        total_days = 0
        for doc in query.stream():
            data = doc.to_dict()
            # حساب عدد الأيام بين تاريخ البداية والنهاية
            start = data.get('startDate')
            end = data.get('endDate')
            if start and end:
                # تحويل التواريخ وحساب الفرق
                from datetime import datetime
                if isinstance(start, str):
                    start = datetime.strptime(start, '%Y-%m-%d')
                if isinstance(end, str):
                    end = datetime.strptime(end, '%Y-%m-%d')
                days = (end - start).days + 1
                total_days += days
        
        return total_days
    except Exception as e:
        print(f"خطأ في جلب طلبات الإجازة للموظف {employee_id}: {e}")
        return 0

# استيراد دوال Firebase من firebase_config
from firebase_config import create_request, get_latest_requests, cancel_request

app = Flask(__name__)
CORS(app)  # Allow static site to call the API

# Configure JSON to handle Arabic text properly
app.config['JSON_AS_ASCII'] = False
app.config['JSONIFY_PRETTYPRINT_REGULAR'] = False  # reduce CPU/bytes in prod

# Enable gzip/deflate compression for JSON and static responses (no-op if missing)
try:
    Compress(app)
except Exception:
    pass

# Lightweight in-process cache (no-op if flask_caching not installed)
try:
    cache = Cache(app, config={
        "CACHE_TYPE": "SimpleCache",
        "CACHE_DEFAULT_TIMEOUT": 30
    })
except Exception:
    cache = Cache()

# Configure file upload limits
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size
app.config['UPLOAD_TIMEOUT'] = 300  # 5 minutes timeout

# JWT Secret Key
JWT_SECRET = os.environ.get('JWT_SECRET', 'your-secret-key-here')

def token_required(f):
    """Decorator للتحقق من صحة JWT token"""
    @wraps(f)
    def decorated(*args, **kwargs):
        token = None
        
        # البحث عن token في header
        if 'Authorization' in request.headers:
            auth_header = request.headers['Authorization']
            try:
                token = auth_header.split(" ")[1]  # Bearer TOKEN
            except IndexError:
                return jsonify({'error': 'Invalid token format'}), 401
        
        if not token:
            return jsonify({'error': 'Token is missing'}), 401
        
        try:
            # فك تشفير token
            data = jwt.decode(token, SECRET, algorithms=['HS256'])
            current_user = data['sub']
        except jwt.ExpiredSignatureError:
            return jsonify({'error': 'Token has expired'}), 401
        except jwt.InvalidTokenError:
            return jsonify({'error': 'Token is invalid'}), 401
        
        return f(current_user, *args, **kwargs)
    
    return decorated

# Helper function for UTF-8 JSON responses
def json_response(data, status_code=200):
    """Create JSON response with proper UTF-8 encoding for Arabic text"""
    response = jsonify(data)
    response.headers['Content-Type'] = 'application/json; charset=utf-8'
    return response, status_code

# Config
SECRET = os.environ.get("APP_SECRET", "CHANGE_ME_SECRET")

# تهيئة Firebase عند بدء التطبيق
firebase_initialized = initialize_firebase()
if not firebase_initialized:
    print("⚠️ تحذير: فشل في تهيئة Firebase، سيتم استخدام البيانات الوهمية")
else:
    # إنشاء المستخدم الافتراضي إذا لم يكن موجوداً
    try:
        admin_user = get_user_by_username("anas")
        if not admin_user:
            admin_data = {
                'username': 'anas',
                'password_hash': generate_password_hash(os.environ.get('DEFAULT_ADMIN_PASSWORD', 'TempPass123!')),
                'is_superadmin': True,
                'services': 'attendance,overtime,employees,stats',
                'is_active': True
            }
            create_user(admin_data)
            print("✅ تم إنشاء المستخدم الافتراضي 'anas'")
        else:
            # تحديث خدمات المستخدم الموجود إذا لم تحتوِ على employees أو stats
            current_services = admin_user.get('services', '')
            services_to_add = []
            
            if 'employees' not in current_services:
                services_to_add.append('employees')
            if 'stats' not in current_services:
                services_to_add.append('stats')
            
            if services_to_add:
                # بناء قائمة الخدمات المحدثة
                services_list = [s.strip() for s in current_services.split(',') if s.strip()]
                services_list.extend(services_to_add)
                updated_services = ','.join(services_list)
                
                from firebase_config import db
                if db:
                    try:
                        users_ref = db.collection('users')
                        query = users_ref.where('username', '==', 'anas').limit(1)
                        docs = list(query.stream())
                        if docs:
                            doc_ref = docs[0].reference
                            doc_ref.update({'services': updated_services})
                            print(f"✅ تم تحديث خدمات المستخدم الافتراضي: {', '.join(services_to_add)}")
                    except Exception as update_error:
                        print(f"⚠️ خطأ في تحديث خدمات المستخدم: {update_error}")
            else:
                print("✅ المستخدم الافتراضي 'anas' موجود بالفعل")
    except Exception as e:
        print(f"⚠️ خطأ في إنشاء المستخدم الافتراضي: {str(e)}")

def create_token(username: str, is_superadmin: bool, services: str) -> str:
    payload = {
        "sub": username,
        "admin": is_superadmin,
        "srv": services,
        "iat": int(datetime.utcnow().timestamp()),
        # expire after 12 hours to minimize abuse/resource leaks from stale tokens
        "exp": int((datetime.utcnow() + timedelta(hours=12)).timestamp()),
    }
    return jwt.encode(payload, SECRET, algorithm="HS256")

def require_auth(required_service: str = None):
    def wrapper(fn):
        def inner(*args, **kwargs):
            auth = request.headers.get("Authorization", "")
            if not auth.startswith("Bearer "):
                return ("Unauthorized", 401)
            token = auth.split(" ", 1)[1]
            try:
                data = jwt.decode(token, SECRET, algorithms=["HS256"])
            except Exception:
                return ("Invalid token", 401)
            request.user = data
            if required_service and (not data.get("admin")):
                srv = (data.get("srv") or "")
                allowed = [s.strip() for s in srv.split(',') if s.strip()]
                if required_service not in allowed:
                    return ("Forbidden", 403)
            return fn(*args, **kwargs)
        inner.__name__ = fn.__name__
        return inner
    return wrapper

# === نقاط النهاية للمصادقة ===

@app.route("/api/login", methods=["POST"])
def login():
    """تسجيل الدخول باستخدام Firebase"""
    try:
        data = request.get_json()
        username = data.get("username", "").strip()
        password = data.get("password", "")
        
        if not username or not password:
            return jsonify({"error": "اسم المستخدم وكلمة المرور مطلوبان"}), 400
        
        user = get_user_by_username(username)
        if not user:
            return jsonify({"error": "اسم المستخدم أو كلمة المرور غير صحيحة"}), 401
        
        # التحقق من كلمة المرور
        if not user or not check_password_hash(user.get('passwordHash', ''), password):
            return jsonify({"error": "اسم المستخدم أو كلمة المرور غير صحيحة"}), 401
        
        # التحقق من حالة الحساب
        if not user.get('is_active', True):
            return jsonify({"error": "تم تعطيل هذا الحساب. تواصل مع الإدارة."}), 403
        
        token = create_token(
            username=user['username'],
            is_superadmin=user.get('isSuperadmin', False),
            services=user.get('services', '')
        )
        
        return jsonify({
            "token": token,
            "username": user['username'],
            "is_superadmin": user.get('isSuperadmin', False),
            "services": user.get('services', '').split(',') if user.get('services') else []
        })
        
    except Exception as e:
        print(f"خطأ في تسجيل الدخول: {str(e)}")
        return jsonify({"error": "خطأ في الخادم"}), 500

@app.route("/api/signup", methods=["POST"])
def signup():
    """طلب إنشاء حساب جديد"""
    try:
        data = request.get_json()
        username = data.get("username", "").strip()
        password = data.get("password", "")
        
        if not username or not password:
            return json_response({"error": "اسم المستخدم وكلمة المرور مطلوبان"}, 400)
        
        if len(password) < 6:
            return json_response({"error": "كلمة المرور يجب أن تكون 6 أحرف على الأقل"}, 400)
        
        # التحقق من عدم وجود المستخدم
        existing_user = get_user_by_username(username)
        if existing_user:
            return json_response({"error": "اسم المستخدم موجود بالفعل"}, 400)
        
        # إضافة طلب معلق
        password_hash = generate_password_hash(password)
        print(f"🔄 محاولة إضافة طلب معلق للمستخدم: {username}")
        success = add_pending_user(username, password_hash)
        
        if success:
            print(f"✅ تم إضافة الطلب المعلق بنجاح: {username}")
            return json_response({"message": "تم إرسال طلبك، في انتظار موافقة المدير"})
        else:
            print(f"❌ فشل في إضافة الطلب المعلق: {username}")
            return json_response({"error": "فشل في إرسال الطلب. تحقق من إعدادات قاعدة البيانات أو اتصل بالمدير"}, 500)
            
    except Exception as e:
        print(f"❌ خطأ في التسجيل: {str(e)}")
        import traceback
        traceback.print_exc()
        return json_response({"error": f"خطأ في الخادم: {str(e)}"}, 500)

# === نقاط النهاية الإدارية ===

@app.route("/api/admin/pending", methods=["GET"])
@require_auth()
def get_pending():
    """جلب طلبات الحسابات المعلقة (للمدراء فقط)"""
    try:
        if not request.user.get("admin"):
            return jsonify({"error": "غير مسموح"}), 403
        
        pending_users = get_pending_users()
        
        # تحويل التواريخ إلى نص
        for user in pending_users:
            if 'createdAt' in user and user['createdAt']:
                user['created_at'] = user['createdAt'].strftime('%Y-%m-%d %H:%M:%S')
        
        return jsonify(pending_users)
        
    except Exception as e:
        print(f"خطأ في جلب الطلبات المعلقة: {str(e)}")
        return jsonify({"error": "خطأ في الخادم"}), 500

@app.route("/api/admin/approve", methods=["POST"])
@require_auth()
def approve_user():
    """الموافقة على طلب حساب"""
    try:
        if not request.user.get("admin"):
            return jsonify({"error": "غير مسموح"}), 403
        
        data = request.get_json()
        username = data.get("username", "").strip()
        services = data.get("services", [])
        
        if not username:
            return jsonify({"error": "اسم المستخدم مطلوب"}), 400
        
        # تحويل services من array إلى string
        services_str = ','.join(services) if isinstance(services, list) else str(services)
        
        success = approve_pending_user(username, services_str)
        
        if success:
            return jsonify({"message": "تم قبول المستخدم بنجاح"})
        else:
            return jsonify({"error": "فشل في قبول المستخدم"}), 500
            
    except Exception as e:
        print(f"خطأ في قبول المستخدم: {str(e)}")
        return jsonify({"error": "خطأ في الخادم"}), 500

@app.route("/api/admin/reject", methods=["POST"])
@require_auth()
def reject_user():
    """رفض طلب حساب"""
    try:
        if not request.user.get("admin"):
            return json_response({"error": "غير مسموح"}, 403)
        
        data = request.get_json()
        username = data.get("username", "").strip()
        
        print(f"🔄 طلب رفض مستخدم: {username}")
        
        if not username:
            return json_response({"error": "اسم المستخدم مطلوب"}, 400)
        
        success = reject_pending_user(username)
        
        if success:
            print(f"✅ تم رفض المستخدم بنجاح: {username}")
            return json_response({"message": "تم رفض المستخدم بنجاح"})
        else:
            print(f"❌ فشل في رفض المستخدم: {username}")
            return json_response({"error": "فشل في رفض المستخدم. قد يكون المستخدم غير موجود"}, 404)
            
    except Exception as e:
        print(f"❌ خطأ في رفض المستخدم: {str(e)}")
        import traceback
        traceback.print_exc()
        return json_response({"error": "خطأ في الخادم"}, 500)

@app.route("/api/admin/delete", methods=["POST"])
@require_auth()
def delete_user_endpoint():
    """حذف مستخدم نهائياً"""
    try:
        if not request.user.get("admin"):
            return json_response({"error": "غير مسموح"}, 403)
        
        data = request.get_json()
        username = data.get("username", "").strip()
        
        print(f"🔄 طلب حذف مستخدم: {username}")
        
        if not username:
            return json_response({"error": "اسم المستخدم مطلوب"}, 400)
        
        # التحقق من عدم حذف المدير الحالي
        if username == request.user.get("username"):
            return json_response({"error": "لا يمكن حذف حسابك الخاص"}, 400)
        
        # التحقق من وجود المستخدم
        existing_user = get_user_by_username(username)
        if not existing_user:
            return json_response({"error": "المستخدم غير موجود"}, 404)
        
        success = delete_user(username)
        
        if success:
            print(f"✅ تم حذف المستخدم بنجاح: {username}")
            return json_response({"message": f"تم حذف المستخدم '{username}' نهائياً"})
        else:
            print(f"❌ فشل في حذف المستخدم: {username}")
            return json_response({"error": "فشل في حذف المستخدم"}, 500)
            
    except Exception as e:
        print(f"❌ خطأ في حذف المستخدم: {str(e)}")
        import traceback
        traceback.print_exc()
        return json_response({"error": "خطأ في الخادم"}, 500)

@app.route("/api/admin/users", methods=["GET"])
@require_auth()
def get_users():
    """جلب جميع المستخدمين (للمدراء فقط)"""
    try:
        if not request.user.get("admin"):
            return jsonify({"error": "غير مسموح"}), 403
        
        users = get_all_users()
        
        # إخفاء كلمات المرور وتحويل التواريخ وتحويل services إلى array
        for user in users:
            user.pop('passwordHash', None)
            if 'createdAt' in user and user['createdAt']:
                user['created_at'] = user['createdAt'].strftime('%Y-%m-%d %H:%M:%S')
            # تحويل services من string إلى array
            if 'services' in user:
                if isinstance(user['services'], str):
                    user['services'] = user['services'].split(',') if user['services'] else []
        
        return jsonify(users)
        
    except Exception as e:
        print(f"خطأ في جلب المستخدمين: {str(e)}")
        return jsonify({"error": "خطأ في الخادم"}), 500

@app.route("/api/admin/users", methods=["POST"])
@require_auth()
def create_user_admin():
    """إنشاء مستخدم جديد (للمدراء فقط)"""
    try:
        if not request.user.get("admin"):
            return jsonify({"error": "غير مسموح"}), 403
        
        data = request.get_json()
        username = data.get("username", "").strip()
        password = data.get("password", "").strip()
        services = data.get("services", [])
        
        if not username or not password:
            return jsonify({"error": "اسم المستخدم وكلمة المرور مطلوبان"}), 400
        
        # تحقق من وجود المستخدم
        existing_user = get_user_by_username(username)
        if existing_user:
            return json_response({"error": "اسم المستخدم موجود بالفعل"}, 400)
        
        # تحويل services إلى string
        services_str = ','.join(services) if isinstance(services, list) else str(services)
        
        # إنشاء المستخدم
        user_data = {
            'username': username,
            'password_hash': generate_password_hash(password),
            'is_superadmin': False,
            'services': services_str,
            'is_active': True
        }
        
        success = create_user(user_data)
        
        if success:
            return jsonify({"message": "تم إنشاء المستخدم بنجاح"})
        else:
            return jsonify({"error": "فشل في إنشاء المستخدم"}), 500
        
    except Exception as e:
        print(f"خطأ في إنشاء المستخدم: {str(e)}")
        return jsonify({"error": "خطأ في الخادم"}), 500

@app.route("/api/admin/toggle-status", methods=["POST"])
@require_auth()
def toggle_user_status():
    """تفعيل/تعطيل حساب مستخدم (للمدراء فقط)"""
    try:
        if not request.user.get("admin"):
            return jsonify({"error": "غير مسموح"}), 403
        
        data = request.get_json()
        username = data.get("username", "").strip()
        
        if not username:
            return jsonify({"error": "اسم المستخدم مطلوب"}), 400
        
        # جلب المستخدم الحالي
        user = get_user_by_username(username)
        if not user:
            return jsonify({"error": "المستخدم غير موجود"}), 404
        
        # تبديل حالة الحساب
        current_status = user.get('is_active', True)
        new_status = not current_status
        
        # تحديث الحساب في قاعدة البيانات
        from firebase_config import get_db
        db = get_db()
        users_ref = db.collection('users')
        query = users_ref.where('username', '==', username).limit(1)
        docs = query.stream()
        
        updated = False
        for doc in docs:
            doc.reference.update({'is_active': new_status})
            updated = True
            break
        
        if updated:
            status_text = "تم تفعيل" if new_status else "تم تعطيل"
            return jsonify({"message": f"{status_text} حساب المستخدم '{username}' بنجاح"})
        else:
            return jsonify({"error": "فشل في تحديث حالة المستخدم"}), 500
        
    except Exception as e:
        print(f"خطأ في تبديل حالة المستخدم: {str(e)}")
        return jsonify({"error": "خطأ في الخادم"}), 500

@app.route("/api/admin/users/update", methods=["POST"])
@require_auth()
def update_user():
    """تحديث بيانات مستخدم (للمدراء فقط)"""
    try:
        if not request.user.get("admin"):
            return jsonify({"error": "غير مسموح"}), 403
        
        data = request.get_json()
        old_username = data.get("old_username", "").strip()  # اسم المستخدم الحالي
        new_username = data.get("username", "").strip()      # اسم المستخدم الجديد
        services = data.get("services", [])
        password = data.get("password", "").strip()
        
        # استخدام old_username للبحث، أو username إذا لم يتم توفير old_username
        search_username = old_username if old_username else new_username
        
        if not search_username:
            return jsonify({"error": "اسم المستخدم مطلوب"}), 400
        
        # جلب المستخدم الحالي
        user = get_user_by_username(search_username)
        if not user:
            return jsonify({"error": "المستخدم غير موجود"}), 404
        
        # تحضير البيانات للتحديث
        update_data = {}
        
        # تحديث اسم المستخدم إذا تغير
        if new_username and new_username != search_username:
            # التحقق من عدم وجود اسم المستخدم الجديد
            existing_user = get_user_by_username(new_username)
            if existing_user:
                return jsonify({"error": f"اسم المستخدم '{new_username}' موجود بالفعل"}), 400
            update_data['username'] = new_username
        
        # تحديث الخدمات
        if services:
            services_str = ','.join(services) if isinstance(services, list) else str(services)
            update_data['services'] = services_str
        
        # تحديث كلمة المرور إذا تم توفيرها
        if password:
            update_data['passwordHash'] = generate_password_hash(password)
        
        if not update_data:
            return jsonify({"error": "لا توجد بيانات للتحديث"}), 400
        
        # تحديث المستخدم في قاعدة البيانات
        from firebase_config import get_db
        db = get_db()
        users_ref = db.collection('users')
        query = users_ref.where('username', '==', search_username).limit(1)
        docs = query.stream()
        
        updated = False
        for doc in docs:
            doc.reference.update(update_data)
            updated = True
            break
        
        if updated:
            final_username = new_username if new_username and new_username != search_username else search_username
            return jsonify({"message": f"تم تحديث بيانات المستخدم '{final_username}' بنجاح"})
        else:
            return jsonify({"error": "فشل في تحديث المستخدم"}), 500
        
    except Exception as e:
        print(f"خطأ في تحديث المستخدم: {str(e)}")
        return jsonify({"error": "خطأ في الخادم"}), 500

# === نقاط النهاية للطلبات ===

@app.route("/api/firebase/status", methods=["GET"])
def firebase_status():
    """فحص حالة اتصال Firebase"""
    try:
        from firebase_config import get_db
        db = get_db()
        
        if not db:
            return jsonify({
                "status": "disconnected",
                "message": "Firebase غير متصل"
            }), 500
        
        # محاولة جلب عدد الطلبات
        try:
            requests_ref = db.collection('requests')
            count = None
            try:
                # استخدام تجميع العد إذا كان مدعوماً من المكتبة
                # ملاحظة: صيغة الوصول قد تختلف حسب إصدار المكتبة
                count_query = requests_ref.count()
                agg = count_query.get()
                # حاول استخراج القيمة من النتيجة (تختلف البنية باختلاف الإصدارات)
                if isinstance(agg, list) and agg:
                    # google-cloud-firestore >= 2.7 يرجع List[AggregationResult]
                    first = agg[0]
                    # بعض الإصدارات تستخدم first[0].value
                    count = getattr(first, 'value', None) or getattr(first[0], 'value', None)
                    if count is None:
                        # fallback أخير
                        count = int(str(first)) if str(first).isdigit() else None
            except Exception:
                count = None
            if count is None:
                #Fallback آمن: عدّ يدوي (أثقل) لكنه يعمل إذا لم تتوفر count()
                count = sum(1 for _ in requests_ref.stream())
            
            return jsonify({
                "status": "connected",
                "message": "Firebase متصل بنجاح",
                "requests_count": int(count)
            })
        except Exception as e:
            print(f"خطأ في الوصول للبيانات: {str(e)}")
            return jsonify({
                "status": "error",
                "message": f"خطأ في الوصول للبيانات: {str(e)}"
            }), 500
            
    except Exception as e:
        return jsonify({
            "status": "error", 
            "message": f"خطأ في Firebase: {str(e)}"
        }), 500

@app.route("/api/requests/reset", methods=["POST"])
@require_auth("overtime")
def reset_all_requests():
    """إعادة تعيين جميع الطلبات لحالة نشط (للاختبار)"""
    try:
        from firebase_config import get_db
        db = get_db()
        
        if not db:
            return jsonify({"error": "Firebase غير متصل"}), 500
        
        # جلب جميع الطلبات
        requests_ref = db.collection('requests')
        all_docs = list(requests_ref.stream())
        
        # تحديث على دفعات لتقليل عدد الرحلات إلى Firestore
        batch = db.batch()
        updated_count = 0
        for i, doc in enumerate(all_docs, start=1):
            batch.update(doc.reference, {
                'status': 'active',
                'canceledBy': None,
                'canceledAt': None
            })
            updated_count += 1
            # نفذ الكوميت كل 400 عملية (حد آمن لدفعة واحدة)
            if i % 400 == 0:
                batch.commit()
                batch = db.batch()
        # كوميت أخير
        batch.commit()
        
        return jsonify({
            "message": f"تم إعادة تعيين {updated_count} طلب لحالة نشط",
            "count": updated_count
        })
        
    except Exception as e:
        print(f"خطأ في إعادة التعيين: {str(e)}")
        return jsonify({"error": "خطأ في الخادم"}), 500

@app.route("/api/requests/test", methods=["POST"])
@require_auth("overtime")
def create_test_request():
    """إنشاء طلب تجريبي للاختبار"""
    try:
        # إنشاء طلب تجريبي
        test_request = {
            'employee_id': '12345',
            'kind': 'overtime',
            'date': '2025-01-01',
            'reason': 'طلب تجريبي للاختبار',
            'supervisor': request.user.get("sub", "test_supervisor")
        }
        
        success = create_request(test_request)
        
        if success:
            return jsonify({"message": "تم إنشاء طلب تجريبي بنجاح"})
        else:
            return jsonify({"error": "فشل في إنشاء الطلب التجريبي"}), 500
            
    except Exception as e:
        print(f"خطأ في إنشاء الطلب التجريبي: {str(e)}")
        return jsonify({"error": "خطأ في الخادم"}), 500

@app.route("/api/requests/create", methods=["POST"])
@require_auth("overtime")
def create_request_endpoint():
    """إنشاء طلب جديد (إضافي/إجازة)"""
    try:
        data = request.get_json()
        employee_id = data.get("employee_id", "").strip()
        kind = data.get("kind", "").strip()
        req_date = data.get("date", "").strip()
        reason = data.get("reason", "").strip()
        
        if not employee_id or not kind or not req_date:
            return jsonify({"error": "معرف الموظف ونوع الطلب والتاريخ مطلوبة"}), 400
        
        if kind not in ["overtime", "leave"]:
            return jsonify({"error": "نوع الطلب يجب أن يكون overtime أو leave"}), 400
        
        # إنشاء الطلب
        request_data = {
            "employee_id": employee_id,
            "kind": kind,
            "date": req_date,
            "reason": reason,
            "supervisor": request.user.get("sub", "")
        }
        
        # إضافة ساعات الإضافي إذا كان النوع overtime
        if kind == "overtime":
            hours = data.get("hours", 0)
            try:
                request_data["hours"] = float(hours)
            except (ValueError, TypeError):
                return jsonify({"error": "ساعات الإضافي يجب أن تكون رقماً"}), 400
        
        # إضافة تاريخ النهاية إذا كان النوع leave
        if kind == "leave":
            end_date = data.get("end_date", req_date).strip()
            request_data["end_date"] = end_date
        
        success = create_request(request_data)
        
        if success:
            return jsonify({"message": "تم إنشاء الطلب بنجاح"})
        else:
            return jsonify({"error": "فشل في إنشاء الطلب"}), 500
            
    except Exception as e:
        print(f"خطأ في إنشاء الطلب: {str(e)}")
        return jsonify({"error": "خطأ في الخادم"}), 500

@app.route("/api/requests/latest", methods=["GET"])
@require_auth("overtime")
def get_latest_requests_endpoint():
    """جلب أحدث الطلبات"""
    try:
        limit = int(request.args.get("limit", 10))
        cache_key = f"latest_requests:{limit}"

        data = cache.get(cache_key)
        if data is None:
            # جلب من Firestore فقط عند الحاجة ثم التخزين في الذاكرة لمدة قصيرة
            requests_list = get_latest_requests(limit)
            cache.set(cache_key, requests_list, timeout=30)
            data = requests_list
        return jsonify(data)
        
    except Exception as e:
        print(f"❌ خطأ في جلب الطلبات: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": "خطأ في الخادم"}), 500

@app.route("/api/requests/cancel", methods=["POST"])
@require_auth("overtime")
def cancel_request_endpoint():
    """إلغاء طلب"""
    try:
        data = request.get_json()
        request_id = data.get("id")
        
        if not request_id:
            return jsonify({"error": "معرف الطلب مطلوب"}), 400
        
        success = cancel_request(request_id, request.user.get("sub", ""))
        
        if success:
            return jsonify({"message": "تم إلغاء الطلب بنجاح"})
        else:
            return jsonify({"error": "فشل في إلغاء الطلب"}), 500
            
    except Exception as e:
        print(f"خطأ في إلغاء الطلب: {str(e)}")
        return jsonify({"error": "خطأ في الخادم"}), 500

@app.route("/api/requests/enable", methods=["POST"])
@require_auth("overtime")
def enable_request_endpoint():
    """تفعيل طلب"""
    try:
        data = request.get_json()
        request_id = data.get("id")
        
        if not request_id:
            return jsonify({"error": "معرف الطلب مطلوب"}), 400
        
        from firebase_config import get_db
        db = get_db()
        
        if not db:
            return jsonify({"error": "Firebase غير متصل"}), 500
        
        # محاولة البحث بـ document ID أولاً
        try:
            doc_ref = db.collection('requests').document(request_id)
            doc = doc_ref.get()
            
            if doc.exists:
                doc_ref.update({
                    'status': 'active',
                    'canceledBy': None,
                    'canceledAt': None
                })
                print(f"✅ تم تفعيل الطلب: {request_id}")
                return jsonify({"message": "تم تفعيل الطلب بنجاح"})
        except:
            pass
            
        # البحث عن الطلب بـ integer ID
        requests_ref = db.collection('requests')
        try:
            query = requests_ref.where('id', '==', int(request_id))
            docs = list(query.stream())
            
            if docs:
                doc_ref = docs[0].reference
                doc_ref.update({
                    'status': 'active',
                    'canceledBy': None,
                    'canceledAt': None
                })
                print(f"✅ تم تفعيل الطلب: {request_id}")
                return jsonify({"message": "تم تفعيل الطلب بنجاح"})
        except ValueError:
            pass
            
        return jsonify({"error": "لم يتم العثور على الطلب"}), 404
            
    except Exception as e:
        print(f"خطأ في تفعيل الطلب: {str(e)}")
        return jsonify({"error": "خطأ في الخادم"}), 500

@app.route("/api/requests/delete", methods=["DELETE"])
@require_auth("overtime")
def delete_request_endpoint():
    """حذف طلب نهائياً"""
    try:
        data = request.get_json()
        request_id = data.get("id")
        
        if not request_id:
            return jsonify({"error": "معرف الطلب مطلوب"}), 400
        
        from firebase_config import get_db
        db = get_db()
        
        if not db:
            return jsonify({"error": "Firebase غير متصل"}), 500
        
        # محاولة البحث بـ document ID أولاً
        try:
            doc_ref = db.collection('requests').document(request_id)
            doc = doc_ref.get()
            
            if doc.exists:
                doc_ref.delete()
                print(f"✅ تم حذف الطلب: {request_id}")
                return jsonify({"message": "تم حذف الطلب بنجاح"})
        except:
            pass
            
        # البحث عن الطلب بـ integer ID
        requests_ref = db.collection('requests')
        try:
            query = requests_ref.where('id', '==', int(request_id))
            docs = list(query.stream())
            
            if docs:
                docs[0].reference.delete()
                print(f"✅ تم حذف الطلب: {request_id}")
                return jsonify({"message": "تم حذف الطلب بنجاح"})
        except ValueError:
            pass
            
        return jsonify({"error": "لم يتم العثور على الطلب"}), 404
            
    except Exception as e:
        print(f"خطأ في حذف الطلب: {str(e)}")
        return jsonify({"error": "خطأ في الخادم"}), 500

# === نقاط النهاية لمعالج الحضور (تبقى كما هي) ===

@app.route("/api/attendance/analyze", methods=["POST"])
@require_auth("attendance")
def analyze_attendance_file():
    """تحليل ملف الحضور وإرجاع معلومات أساسية"""
    try:
        print(f"🔍 استقبال طلب تحليل ملف الحضور من {request.remote_addr}")
        
        if "file" not in request.files:
            return jsonify({"error": "لم يتم رفع أي ملف"}), 400
        
        file = request.files["file"]
        if file.filename == "":
            return jsonify({"error": "لم يتم اختيار ملف"}), 400
        
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({"error": "نوع الملف غير مدعوم. يرجى رفع ملف Excel"}), 400
        
        # حفظ الملف مؤقتاً
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            file.save(temp_file.name)
            temp_path = temp_file.name
        
        try:
            # تحليل الملف
            sheet_name = request.form.get("sheet", None) or None
            print(f"🔍 بدء تحليل الملف: {file.filename}, الورقة: {sheet_name}")
            
            from attendance_processor import analyze_file
            
            analysis_result = analyze_file(temp_path, sheet_name)
            
            # إضافة معلومات إضافية
            analysis_result["file_name"] = file.filename
            analysis_result["file_size"] = os.path.getsize(temp_path)
            
            print(f"✅ تم تحليل الملف بنجاح:")
            print(f"   - عدد الموظفين: {analysis_result.get('employees_count', 0)}")
            print(f"   - نوع الملف: {analysis_result.get('file_format', 'unknown')}")
            print(f"   - أول تاريخ: {analysis_result.get('first_date', 'N/A')}")
            print(f"   - آخر تاريخ: {analysis_result.get('last_date', 'N/A')}")
            print(f"   - عدد الأيام: {analysis_result.get('period_days', 0)}")
            print(f"   - طلبات الإضافي: {analysis_result.get('overtime_requests_count', 0)}")
            print(f"   - طلبات الإجازة: {analysis_result.get('leave_requests_count', 0)}")
            
            return jsonify({
                "success": True,
                "analysis": analysis_result,
                "message": "تم تحليل الملف بنجاح"
            })
            
        finally:
            # حذف الملف المؤقت
            try:
                os.unlink(temp_path)
            except:
                pass
                
    except Exception as e:
        print(f"❌ خطأ في تحليل الملف: {e}")
        return jsonify({
            "success": False,
            "error": f"خطأ في تحليل الملف: {str(e)}"
        }), 500


@app.route("/api/attendance/sync-employees", methods=["POST"])
@require_auth("attendance")
def sync_employees_from_file():
    """مزامنة الموظفين من ملف الحضور فقط - عملية سريعة"""
    try:
        print(f"🔄 استقبال طلب مزامنة الموظفين من {request.remote_addr}")
        
        if 'file' not in request.files:
            return jsonify({"error": "لم يتم رفع أي ملف"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "لم يتم اختيار ملف"}), 400
        
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({"error": "نوع الملف غير مدعوم. يرجى رفع ملف Excel"}), 400
        
        # حفظ الملف مؤقتاً
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            file.save(tmp_file.name)
            temp_file_path = tmp_file.name
        
        try:
            print(f"📋 معلومات الطلب: Content-Length: {request.content_length}")
            
            # التحقق من حجم الملف لتجنب المعالجة المفرطة
            file_size_mb = request.content_length / (1024 * 1024) if request.content_length else 0
            if file_size_mb > 50:  # ملفات أكبر من 50 ميجابايت
                return jsonify({
                    "error": f"الملف كبير جداً ({file_size_mb:.1f} MB). الحد الأقصى المسموح: 50 MB",
                    "suggestion": "يرجى تقسيم الملف أو استخدام ملف أصغر"
                }), 400
            
            # استخراج بيانات الموظفين فقط
            from attendance_processor import extract_employees_from_file
            employees_data = extract_employees_from_file(temp_file_path)
            
            if not employees_data:
                return jsonify({"error": "لم يتم العثور على بيانات موظفين في الملف"}), 400
            
            print(f"👥 تم العثور على {len(employees_data)} موظف في الملف")
            
            # تحديد ما إذا كان يجب المزامنة بناءً على عدد الموظفين
            if len(employees_data) > 100:
                return jsonify({
                    "warning": f"عدد الموظفين كبير ({len(employees_data)}). يُنصح بالمزامنة اليدوية",
                    "employees_count": len(employees_data),
                    "suggestion": "استخدم خيار 'تعطيل المزامنة التلقائية' لتجنب التأخير"
                }), 202
            
            # مزامنة الموظفين مع قاعدة البيانات
            from firebase_config import sync_employees_batch
            sync_results = sync_employees_batch(employees_data)
            
            if "error" in sync_results:
                return jsonify({"error": sync_results["error"]}), 500
            
            return jsonify({
                "message": "تم مزامنة الموظفين بنجاح",
                "employees_count": len(employees_data),
                "sync_stats": sync_results,
                "employees_preview": employees_data[:5]  # أول 5 موظفين للمعاينة
            })
            
        finally:
            # حذف الملف المؤقت
            try:
                os.unlink(temp_file_path)
            except:
                pass
        
    except Exception as e:
        print(f"❌ خطأ في مزامنة الموظفين: {str(e)}")
        return jsonify({"error": f"خطأ في مزامنة الموظفين: {str(e)}"}), 500


@app.route("/api/attendance/process", methods=["POST"])
@require_auth("attendance")
def process_attendance():
    """معالجة ملف الحضور"""
    try:
        print(f"🔄 استقبال طلب معالجة الحضور من {request.remote_addr}")
        print(f"📋 معلومات الطلب: Content-Length: {request.content_length}")
        
        if "file" not in request.files:
            print("❌ لم يتم العثور على ملف في الطلب")
            return jsonify({"error": "لم يتم رفع أي ملف"}), 400
        
        file = request.files["file"]
        if file.filename == "":
            return jsonify({"error": "لم يتم اختيار ملف"}), 400
        
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({"error": "نوع الملف غير مدعوم. يرجى رفع ملف Excel"}), 400
        
        # حفظ الملف مؤقتاً
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            file.save(temp_file.name)
            temp_path = temp_file.name
        
        try:
            # معالجة الملف - جمع جميع المعاملات
            sheet_name = request.form.get("sheet", None) or None
            target_days = int(request.form.get("target_days", 26))
            holidays_str = request.form.get("holidays", "")
            holidays = parse_holidays(holidays_str) if holidays_str else set()
            special_days_str = request.form.get("special_days", "")
            special_days = parse_holidays(special_days_str) if special_days_str else set()
            cutoff_hour = int(request.form.get("cutoff_hour", 7))
            fmt = request.form.get("format", "auto")
            allow_negative = request.form.get("allow_negative", "0") == "1"
            language = request.form.get("language", "ar")
            
            # خيارات نوع الملفات المطلوبة
            include_summary = request.form.get("include_summary", "1") == "1"
            include_daily = request.form.get("include_daily", "1") == "1"
            
            print(f"📋 خيارات الملفات المطلوبة:")
            print(f"   - تقرير الملخص: {'نعم' if include_summary else 'لا'}")
            print(f"   - التفاصيل اليومية: {'نعم' if include_daily else 'لا'}")
            
            # التحقق من أن المستخدم اختار نوع ملف واحد على الأقل
            if not include_summary and not include_daily:
                return jsonify({"error": "يجب اختيار نوع ملف واحد على الأقل (ملخص أو يومي)"}), 400
            
            # تشخيص الملف قبل المعالجة
            try:
                from openpyxl import load_workbook
                wb = load_workbook(temp_path, data_only=True, read_only=True)
                ws = wb[sheet_name] if sheet_name else wb.worksheets[0]
                
                print(f"تشخيص الملف:")
                print(f"- اسم الورقة: {ws.title}")
                print(f"- عدد الصفوف: {ws.max_row}")
                print(f"- عدد الأعمدة: {ws.max_column}")
                
                # البحث عن "Employee ID:" في أول 20 صف
                employee_found = False
                print("- فحص أول 10 صفوف:")
                for row_num in range(1, min(11, ws.max_row + 1)):
                    cell_value = ws.cell(row=row_num, column=1).value
                    print(f"  الصف {row_num}: '{cell_value}'")
                    if cell_value and "Employee ID:" in str(cell_value):
                        print(f"- ✅ وُجد 'Employee ID:' في الصف {row_num}: {cell_value}")
                        employee_found = True
                        break
                
                if not employee_found:
                    print("- ⚠️ تحذير: لم يتم العثور على 'Employee ID:' في أول 10 صفوف")
                    print("- 💡 تأكد من أن الملف يحتوي على 'Employee ID:' في العمود الأول")
                    
            except Exception as e:
                print(f"خطأ في تشخيص الملف: {e}")
            
            # استدعاء دالة المعالجة بجميع المعاملات
            print(f"🔄 بدء معالجة الملف: {temp_path}")
            print(f"📋 المعاملات:")
            print(f"   - sheet: {sheet_name}")
            print(f"   - target_days: {target_days}")
            print(f"   - holidays: {holidays}")
            print(f"   - special_days: {special_days}")
            print(f"   - format: {fmt}")
            print(f"   - cutoff_hour: {cutoff_hour}")
            
            try:
                summary_results, daily_results = process_workbook(
                    path=temp_path,
                    sheet_name=sheet_name,
                    target_days=target_days,
                    holidays=holidays,
                    special_days=special_days,
                    fmt=fmt,
                    cutoff_hour=cutoff_hour,
                    dup_threshold_minutes=60,
                    assume_missing_exit_hours=5.0
                )
                print(f"✅ تمت المعالجة بنجاح")
            except Exception as processing_error:
                print(f"❌ خطأ في المعالجة: {processing_error}")
                import traceback
                traceback.print_exc()
                return jsonify({"error": f"خطأ في معالجة الملف: {str(processing_error)}"}), 500
            
            print(f"النتائج: summary={len(summary_results)}, daily={len(daily_results)}")
            if summary_results:
                print(f"أول نتيجة: {summary_results[0]}")
            if daily_results:
                print(f"أول تفصيل يومي: {daily_results[0]}")
            
            # التحقق من وجود نتائج
            if not summary_results and not daily_results:
                print("⚠️ لم يتم العثور على نتائج - المعالجة فشلت")
                return jsonify({"error": "لم يتم العثور على بيانات صالحة في الملف"}), 400
            
            # مزامنة بيانات الموظفين مع قاعدة البيانات (إضافة الجدد فقط)
            print("🔄 بدء إضافة الموظفين الجدد فقط (الأولوية لإدارة الموظفين)...")
            new_employees_added = 0
            existing_employees_skipped = 0
            try:
                from firebase_config import sync_employee_from_attendance
                
                for employee_data in summary_results:
                    employee_id = str(employee_data.get('EmployeeID', ''))
                    name = employee_data.get('Name', '')
                    department = employee_data.get('Department', '')
                    
                    if employee_id and name and department:
                        # التحقق من وجود الموظف أولاً
                        from firebase_config import db
                        existing_query = db.collection('employees').where('employee_id', '==', employee_id).limit(1)
                        existing_docs = list(existing_query.stream())
                        
                        if existing_docs:
                            existing_employees_skipped += 1
                        else:
                            if sync_employee_from_attendance(employee_id, name, department):
                                new_employees_added += 1
                
                print(f"✅ تم إضافة {new_employees_added} موظف جديد")
                print(f"⏭️ تم تجاهل {existing_employees_skipped} موظف موجود (الأولوية لإدارة الموظفين)")
                
            except Exception as sync_error:
                print(f"⚠️ خطأ في مزامنة الموظفين: {sync_error}")
                # لا نوقف المعالجة بسبب خطأ المزامنة
            
            # إنشاء ملف ZIP يحتوي على التقارير المطلوبة فقط
            files_to_create = []
            if include_summary:
                files_to_create.append("summary")
            if include_daily:
                files_to_create.append("daily")
            
            print(f"📦 إنشاء ملف ZIP مع الملفات المطلوبة: {', '.join(files_to_create)}")
            print(f"📊 البيانات: {len(summary_results)} موظف، {len(daily_results)} سجل يومي")
            
            zip_buffer = io.BytesIO()
            
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                
                # إنشاء ملف الملخص إذا كان مطلوباً
                if include_summary:
                    print("📊 إنشاء ملف الملخص...")
                    summary_wb = Workbook()
                    summary_ws = summary_wb.active
                    summary_ws.title = get_translation(language, 'summary_title')
                    
                    # إضافة عناوين الملخص
                    summary_headers = get_translation(language, 'summary_headers')
                    for col, header in enumerate(summary_headers, 1):
                        summary_ws.cell(row=1, column=col, value=header)
                    
                    # إضافة بيانات الملخص
                    if summary_results:
                        for row, result in enumerate(summary_results, 2):
                            employee_id = result.get('EmployeeID', '')
                            
                            # الترتيب الجديد للأعمدة
                            summary_ws.cell(row=row, column=1, value=employee_id)                                    # Employee ID
                            summary_ws.cell(row=row, column=2, value=result.get('Name', ''))                        # Employee Name
                            summary_ws.cell(row=row, column=3, value=result.get('Department', ''))                  # Department
                            summary_ws.cell(row=row, column=4, value=result.get('WorkDays', 0))                     # Work Days
                            summary_ws.cell(row=row, column=5, value=result.get('AbsentDays', 0))                   # Absent Days
                            summary_ws.cell(row=row, column=6, value=result.get('WorkedOnHolidays', 0))             # Worked on Holidays
                            summary_ws.cell(row=row, column=7, value=result.get('ExtraDays', 0))                    # Extra Days
                            summary_ws.cell(row=row, column=8, value=round(result.get('TotalHours', 0), 2))         # Total Hours
                            summary_ws.cell(row=row, column=9, value=round(result.get('OvertimeHours', 0), 2))      # Overtime Hours
                            summary_ws.cell(row=row, column=10, value=round(result.get('RequestedOvertimeHours', 0), 2))  # Requested Overtime Hours
                            summary_ws.cell(row=row, column=11, value=round(result.get('DelayHours', 0), 2))        # Delay Hours
                            summary_ws.cell(row=row, column=12, value=result.get('OvertimeRequestsCount', 0))       # Overtime Requests Count
                            summary_ws.cell(row=row, column=13, value=result.get('LeaveRequestsCount', 0))          # Leave Requests Count
                            summary_ws.cell(row=row, column=14, value=result.get('AssumedExitDays', 0))             # Missing Punches
                    else:
                        # إضافة رسالة عدم وجود بيانات
                        summary_ws.cell(row=2, column=1, value=get_translation(language, 'no_data'))
                        summary_ws.cell(row=2, column=2, value=get_translation(language, 'check_format'))
                    
                    # حفظ ملف الملخص في الذاكرة
                    summary_buffer = io.BytesIO()
                    summary_wb.save(summary_buffer)
                    summary_buffer.seek(0)
                    zip_file.writestr(get_translation(language, 'summary_filename'), summary_buffer.getvalue())
                    print(f"✅ تم إنشاء ملف الملخص مع {len(summary_results)} موظف")
                
                # إنشاء ملف التفاصيل اليومية إذا كان مطلوباً
                if include_daily:
                    daily_wb = Workbook()
                    daily_ws = daily_wb.active
                    daily_ws.title = get_translation(language, 'daily_title')
                    
                    # إضافة عناوين التفاصيل اليومية
                    daily_headers = get_translation(language, 'daily_headers')
                    for col, header in enumerate(daily_headers, 1):
                        daily_ws.cell(row=1, column=col, value=header)
                    
                    # إضافة بيانات التفاصيل اليومية
                    if daily_results:
                        for row, daily in enumerate(daily_results, 2):
                            # استخراج أول وآخر وقت من TimesList
                            times_list = daily.get('TimesList', '')
                            first_in = ''
                            last_out = ''
                            if times_list:
                                times = times_list.split(',')
                                if len(times) >= 1:
                                    first_in = times[0]
                                if len(times) >= 2:
                                    last_out = times[-1]
                            
                            daily_ws.cell(row=row, column=1, value=daily.get('EmployeeID', ''))
                            daily_ws.cell(row=row, column=2, value=daily.get('Name', ''))
                            daily_ws.cell(row=row, column=3, value=daily.get('Department', ''))
                            daily_ws.cell(row=row, column=4, value=str(daily.get('Date', '')))
                            daily_ws.cell(row=row, column=5, value=first_in)
                            daily_ws.cell(row=row, column=6, value=last_out)
                            daily_ws.cell(row=row, column=7, value=round(daily.get('DayHours', 0), 2))
                            daily_ws.cell(row=row, column=8, value=round(daily.get('DayOvertimeHours', 0), 2))
                            daily_ws.cell(row=row, column=9, value=round(daily.get('DayDelayHours', 0), 2))
                            daily_ws.cell(row=row, column=10, value=daily.get('TimesCount', 0))
                            daily_ws.cell(row=row, column=11, value=get_translation(language, 'yes') if daily.get('IsHoliday', 0) == 1 else get_translation(language, 'no'))
                            # إضافة معلومات الطلبات
                            daily_ws.cell(row=row, column=12, value=get_translation(language, 'yes') if daily.get('HasOvertimeRequest', False) else get_translation(language, 'no'))
                            daily_ws.cell(row=row, column=13, value=get_translation(language, 'yes') if daily.get('HasLeaveRequest', False) else get_translation(language, 'no'))
                            daily_ws.cell(row=row, column=14, value=daily.get('OvertimeRequestReason', ''))
                            daily_ws.cell(row=row, column=15, value=daily.get('LeaveRequestReason', ''))
                    else:
                        # إضافة رسالة عدم وجود بيانات
                        daily_ws.cell(row=2, column=1, value=get_translation(language, 'no_daily_data'))
                        daily_ws.cell(row=2, column=2, value=get_translation(language, 'check_format'))
                    
                    # حفظ ملف التفاصيل في الذاكرة
                    daily_buffer = io.BytesIO()
                    daily_wb.save(daily_buffer)
                    daily_buffer.seek(0)
                    zip_file.writestr(get_translation(language, 'daily_filename'), daily_buffer.getvalue())
                    print(f"✅ تم إنشاء ملف التفاصيل اليومية مع {len(daily_results)} سجل")
            
            zip_buffer.seek(0)
            
            # طباعة ملخص الملفات المُنشأة
            created_files = []
            if include_summary:
                created_files.append("ملف الملخص")
            if include_daily:
                created_files.append("ملف التفاصيل اليومية")
            
            print(f"📦 تم إنشاء ملف ZIP يحتوي على: {', '.join(created_files)}")
            
            # إرسال ملف ZIP
            zip_filename = f"{get_translation(language, 'zip_filename')}_{datetime.now().strftime('%Y%m%dT%H%M%S')}.zip"
            return send_file(
                zip_buffer,
                as_attachment=True,
                download_name=zip_filename,
                mimetype='application/zip'
            )
            
        finally:
            # تنظيف الملفات المؤقتة
            try:
                os.unlink(temp_path)
            except:
                pass
                
    except Exception as e:
        error_msg = str(e)
        print(f"❌ خطأ في معالجة الحضور: {error_msg}")
        
        # معالجة أخطاء محددة
        if "413" in error_msg or "Request Entity Too Large" in error_msg:
            return jsonify({"error": "الملف كبير جداً. الحد الأقصى 50 ميجابايت."}), 413
        elif "timeout" in error_msg.lower():
            return jsonify({"error": "انتهت مهلة المعالجة. جرب ملف أصغر."}), 408
        elif "connection" in error_msg.lower():
            return jsonify({"error": "مشكلة في الاتصال. تأكد من استقرار الإنترنت."}), 503
        else:
            return jsonify({"error": f"خطأ في معالجة الملف: {error_msg}"}), 500

# === إدارة الموظفين ===

@app.route("/api/employees", methods=["GET"])
@token_required
def get_employees(current_user):
    """جلب قائمة الموظفين"""
    try:
        from firebase_config import get_all_employees
        employees = get_all_employees()
        return jsonify(employees)
    except Exception as e:
        print(f"خطأ في جلب الموظفين: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/employees", methods=["POST"])
@token_required
def create_employee(current_user):
    """إنشاء موظف جديد"""
    try:
        from firebase_config import db, create_employee as create_emp
        
        data = request.get_json()
        
        # التحقق من البيانات المطلوبة
        required_fields = ['employee_id', 'name', 'department']
        for field in required_fields:
            if not data.get(field):
                return jsonify({"error": f"الحقل {field} مطلوب"}), 400
        
        # التحقق من عدم وجود موظف بنفس الرقم
        employee_id = data['employee_id']
        emp_ref = db.collection('employees').document(employee_id)
        if emp_ref.get().exists:
            return jsonify({"error": f"موظف برقم {employee_id} موجود بالفعل"}), 400
        
        created_id = create_emp(data)
        
        return jsonify({
            "message": "تم إنشاء الموظف بنجاح",
            "id": created_id
        }), 201
        
    except Exception as e:
        print(f"خطأ في إنشاء الموظف: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/employees/<employee_id>", methods=["GET"])
@token_required
def get_employee(current_user, employee_id):
    """جلب بيانات موظف محدد"""
    try:
        from firebase_config import get_employee_by_id
        employee = get_employee_by_id(employee_id)
        
        if not employee:
            return jsonify({"error": "الموظف غير موجود"}), 404
            
        return jsonify(employee)
        
    except Exception as e:
        print(f"خطأ في جلب بيانات الموظف: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/employees/<employee_id>", methods=["PUT"])
@token_required
def update_employee(current_user, employee_id):
    """تحديث بيانات موظف"""
    try:
        data = request.get_json()
        
        from firebase_config import update_employee as update_emp
        success = update_emp(employee_id, data)
        
        if not success:
            return jsonify({"error": "فشل في تحديث الموظف"}), 400
            
        return jsonify({"message": "تم تحديث الموظف بنجاح"})
        
    except Exception as e:
        print(f"خطأ في تحديث الموظف: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/employees/<employee_id>", methods=["DELETE"])
@token_required
def delete_employee(current_user, employee_id):
    """حذف موظف"""
    try:
        from firebase_config import delete_employee as delete_emp
        success = delete_emp(employee_id)
        
        if not success:
            return jsonify({"error": "فشل في حذف الموظف"}), 400
            
        return jsonify({"message": "تم حذف الموظف بنجاح"})
        
    except Exception as e:
        print(f"خطأ في حذف الموظف: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/employees/cleanup-duplicates", methods=["POST"])
@token_required
def cleanup_duplicate_employees(current_user):
    """تنظيف الموظفين المكررين - الاحتفاظ بالأحدث"""
    try:
        print(f"🧹 بدء تنظيف الموظفين المكررين بواسطة {current_user}")
        
        from firebase_config import db
        from datetime import datetime
        
        # جلب جميع الموظفين
        employees_ref = db.collection('employees')
        docs = employees_ref.stream()
        
        # تجميع الموظفين حسب employee_id
        employees_by_id = {}
        all_docs = []
        
        for doc in docs:
            doc_data = doc.to_dict()
            emp_id = doc_data.get('employee_id') or doc_data.get('id')
            
            # تجاهل الوثائق بدون employee_id صحيح
            if not emp_id:
                print(f"⚠️ وثيقة بدون employee_id: {doc.id}")
                continue
                
            emp_id = str(emp_id).strip()
            
            if emp_id not in employees_by_id:
                employees_by_id[emp_id] = []
            
            doc_info = {
                'doc_id': doc.id,
                'data': doc_data,
                'updated_at': doc_data.get('updated_at', ''),
                'created_at': doc_data.get('created_at', ''),
                'employee_id': emp_id
            }
            
            employees_by_id[emp_id].append(doc_info)
            all_docs.append(doc_info)
        
        # البحث عن المكررات
        duplicates_found = 0
        duplicates_removed = 0
        batch = db.batch()
        batch_operations = 0
        
        for emp_id, employee_docs in employees_by_id.items():
            if len(employee_docs) > 1:
                duplicates_found += len(employee_docs) - 1
                print(f"🔍 وجد {len(employee_docs)} نسخ للموظف {emp_id}")
                
                # البحث عن الوثيقة التي معرفها = employee_id (الصحيحة)
                correct_doc = None
                wrong_docs = []
                
                for doc_info in employee_docs:
                    if doc_info['doc_id'] == emp_id:
                        correct_doc = doc_info
                        print(f"  ✅ وجدت الوثيقة الصحيحة: {doc_info['doc_id']} (معرف الوثيقة = رقم الموظف)")
                    else:
                        wrong_docs.append(doc_info)
                
                # إذا لم توجد وثيقة صحيحة، اختر الأحدث وانقلها للمعرف الصحيح
                if not correct_doc:
                    # ترتيب حسب تاريخ التحديث (الأحدث أولاً)
                    employee_docs.sort(key=lambda x: x.get('updated_at', ''), reverse=True)
                    newest_doc = employee_docs[0]
                    
                    print(f"  📝 إنشاء وثيقة جديدة بالمعرف الصحيح: {emp_id}")
                    
                    # إنشاء وثيقة جديدة بالمعرف الصحيح
                    correct_ref = db.collection('employees').document(emp_id)
                    new_data = newest_doc['data'].copy()
                    new_data['employee_id'] = emp_id
                    new_data['id'] = emp_id
                    new_data['updated_at'] = datetime.now().isoformat()
                    
                    batch.set(correct_ref, new_data)
                    batch_operations += 1
                    
                    # حذف جميع الوثائق الخاطئة
                    wrong_docs = employee_docs
                else:
                    # الاحتفاظ بالوثيقة الصحيحة وحذف الباقي
                    pass
                
                # حذف جميع الوثائق الخاطئة
                for wrong_doc in wrong_docs:
                    doc_ref = db.collection('employees').document(wrong_doc['doc_id'])
                    batch.delete(doc_ref)
                    batch_operations += 1
                    duplicates_removed += 1
                    print(f"  🗑️ حذف المكرر: {wrong_doc['doc_id']}")
                    
                    # تنفيذ batch كل 100 عملية
                    if batch_operations >= 100:
                        print(f"💾 حفظ دفعة من {batch_operations} عملية حذف...")
                        batch.commit()
                        batch = db.batch()
                        batch_operations = 0
        
        # تنفيذ آخر batch
        if batch_operations > 0:
            print(f"💾 حفظ الدفعة الأخيرة ({batch_operations} عملية)...")
            batch.commit()
        
        result = {
            "success": True,
            "duplicates_found": duplicates_found,
            "duplicates_removed": duplicates_removed,
            "message": f"تم تنظيف {duplicates_removed} موظف مكرر من أصل {duplicates_found} مكرر موجود"
        }
        
        print(f"✅ اكتمل تنظيف المكررات: {duplicates_removed} حذف")
        return jsonify(result)
        
    except Exception as e:
        print(f"❌ خطأ في تنظيف المكررات: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"خطأ في تنظيف المكررات: {str(e)}"}), 500

@app.route("/api/employees/bulk-delete", methods=["POST"])
@token_required
def bulk_delete_employees(current_user):
    """حذف متعدد للموظفين"""
    try:
        print(f"🗑️ استقبال طلب حذف متعدد من {current_user}")
        
        from firebase_config import db
        
        data = request.get_json()
        if not data:
            print("❌ لم يتم استقبال بيانات JSON")
            return jsonify({"error": "لم يتم إرسال بيانات"}), 400
            
        employee_ids = data.get('employee_ids', [])
        print(f"📋 الموظفين المطلوب حذفهم: {employee_ids}")
        
        if not employee_ids:
            return jsonify({"error": "لم يتم تحديد موظفين للحذف"}), 400
        
        if len(employee_ids) > 100:
            return jsonify({"error": "لا يمكن حذف أكثر من 100 موظف في المرة الواحدة"}), 400
        
        print(f"🗑️ طلب حذف متعدد من {current_user}: {len(employee_ids)} موظف")
        
        # استخدام batch للحذف المتعدد (أكثر كفاءة)
        batch = db.batch()
        deleted_count = 0
        errors = []
        
        for employee_id in employee_ids:
            try:
                emp_ref = db.collection('employees').document(employee_id)
                
                # التحقق من وجود الموظف قبل الحذف
                emp_doc = emp_ref.get()
                if emp_doc.exists:
                    batch.delete(emp_ref)
                    deleted_count += 1
                    print(f"✅ تم تحضير حذف الموظف: {employee_id}")
                else:
                    errors.append(f"الموظف {employee_id} غير موجود")
                    print(f"⚠️ الموظف {employee_id} غير موجود")
                    
            except Exception as e:
                error_msg = f"خطأ في تحضير حذف الموظف {employee_id}: {str(e)}"
                errors.append(error_msg)
                print(f"❌ {error_msg}")
        
        # تنفيذ الحذف المتعدد
        if deleted_count > 0:
            batch.commit()
            print(f"✅ تم حذف {deleted_count} موظف بنجاح")
        
        result = {
            "success": True,
            "deleted_count": deleted_count,
            "total_requested": len(employee_ids),
            "errors": errors
        }
        
        if errors:
            result["message"] = f"تم حذف {deleted_count} موظف مع {len(errors)} خطأ"
        else:
            result["message"] = f"تم حذف {deleted_count} موظف بنجاح"
        
        return jsonify(result)
        
    except Exception as e:
        print(f"❌ خطأ في الحذف المتعدد: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"خطأ في الخادم: {str(e)}"}), 500

@app.route("/api/employees/<employee_id>/toggle", methods=["POST"])
@token_required
def toggle_employee_status(current_user, employee_id):
    """تفعيل/تعطيل موظف"""
    try:
        data = request.get_json()
        active = data.get('active', True)
        
        from firebase_config import toggle_employee_status as toggle_emp
        success = toggle_emp(employee_id, active)
        
        if not success:
            return jsonify({"error": "فشل في تغيير حالة الموظف"}), 400
            
        status_text = "تفعيل" if active else "تعطيل"
        return jsonify({"message": f"تم {status_text} الموظف بنجاح"})
        
    except Exception as e:
        print(f"خطأ في تغيير حالة الموظف: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/user/add-service", methods=["POST"])
@token_required
def add_user_service(current_user):
    """إضافة خدمة لمستخدم"""
    try:
        data = request.get_json()
        service_name = data.get('service', '')
        
        if not service_name:
            return jsonify({"error": "اسم الخدمة مطلوب"}), 400
        
        # جلب بيانات المستخدم الحالي
        from firebase_config import get_user_by_username, db
        user_data = get_user_by_username(current_user)
        
        if not user_data:
            return jsonify({"error": "المستخدم غير موجود"}), 404
        
        # تحديث خدمات المستخدم
        current_services = user_data.get('services', '')
        services_list = [s.strip() for s in current_services.split(',') if s.strip()]
        
        if service_name not in services_list:
            services_list.append(service_name)
            updated_services = ','.join(services_list)
            
            # تحديث في قاعدة البيانات
            if db:
                users_ref = db.collection('users')
                query = users_ref.where('username', '==', current_user).limit(1)
                docs = list(query.stream())
                if docs:
                    doc_ref = docs[0].reference
                    doc_ref.update({'services': updated_services})
                    
                    return jsonify({
                        "message": f"تم إضافة خدمة {service_name} بنجاح",
                        "services": updated_services
                    })
        
        return jsonify({"message": "الخدمة موجودة بالفعل"})
        
    except Exception as e:
        print(f"خطأ في إضافة الخدمة: {e}")
        return jsonify({"error": str(e)}), 500

# === نقاط النهاية للإحصائيات ===

@app.route("/api/stats/dashboard", methods=["GET"])
@require_auth("stats")
def get_dashboard_stats():
    """جلب إحصائيات سريعة للوحة التحكم - يتطلب صلاحية stats"""
    try:
        from firebase_config import get_db, get_all_employees
        
        # استخدم كاش لحماية الحصة وتقليل زمن الاستجابة
        cache_key = "stats:dashboard"
        cached = cache.get(cache_key)
        if cached is not None:
            return jsonify(cached)

        # إحصائيات الموظفين
        total_employees = 0
        active_employees = 0
        try:
            employees = get_all_employees()
            total_employees = len(employees)
            active_employees = len([e for e in employees if e.get('active', True)])
        except ResourceExhausted:
            # في حال نفاد الحصة نُعيد ما يتوفر فقط بدون تحميل كامل
            pass

        # إحصائيات الطلبات
        db = get_db()
        stats = {
            "employees": {
                "total": total_employees,
                "active": active_employees,
                "inactive": total_employees - active_employees
            },
            "requests": {
                "total": 0,
                "active": 0,
                "cancelled": 0,
                "overtime": 0,
                "leave": 0
            },
            "users": {
                "total": 0,
                "pending": 0
            }
        }
        
        if db:
            try:
                # إجمالي الطلبات عبر aggregation count إن توفرت
                requests_ref = db.collection('requests')
                try:
                    stats["requests"]["total"] = int(requests_ref.count().get()[0][0].value)
                except Exception:
                    #Fallback آمن: حد أقصى 1000 لتجنب OOM
                    stats["requests"]["total"] = sum(1 for _ in requests_ref.limit(1000).stream())

                # احصائيات نوع/حالة عبر عينة أحدث 100 فقط لتقليل الكلفة
                latest = requests_ref.order_by('createdAt', direction='DESCENDING').limit(100).stream()
                for req in latest:
                    rq = req.to_dict() or {}
                    status = rq.get('status', 'active')
                    kind = rq.get('kind') or rq.get('type', '')
                    if status in ('cancelled', 'canceled'):
                        stats["requests"]["cancelled"] += 1
                    else:
                        stats["requests"]["active"] += 1
                    if kind == 'overtime':
                        stats["requests"]["overtime"] += 1
                    elif kind == 'leave':
                        stats["requests"]["leave"] += 1
            except ResourceExhausted:
                pass

            # إحصائيات المستخدمين (count aggregation أو حد أعلى)
            try:
                users_ref = db.collection('users')
                pending_ref = db.collection('pendingUsers')
                try:
                    stats["users"]["total"] = int(users_ref.count().get()[0][0].value)
                except Exception:
                    stats["users"]["total"] = sum(1 for _ in users_ref.limit(1000).stream())
                try:
                    stats["users"]["pending"] = int(pending_ref.count().get()[0][0].value)
                except Exception:
                    stats["users"]["pending"] = sum(1 for _ in pending_ref.limit(1000).stream())
            except ResourceExhausted:
                pass
        
        # خزّن 60 ثانية
        cache.set(cache_key, stats, timeout=60)
        return jsonify(stats)
        
    except Exception as e:
        print(f"❌ خطأ في جلب الإحصائيات: {str(e)}")
        return jsonify({"error": "خطأ في جلب الإحصائيات"}), 500

@app.route("/api/stats/recent-activity", methods=["GET"])
@require_auth("stats")
def get_recent_activity():
    """جلب آخر الأنشطة في النظام - يتطلب صلاحية stats"""
    try:
        from firebase_config import get_db
        
        # كاش قصير المدى 30 ثانية
        cache_key = "stats:recent"
        cached = cache.get(cache_key)
        if cached is not None:
            return jsonify(cached)

        db = get_db()
        activities = []
        
        if db:
            try:
                # جلب آخر 10 طلبات فقط
                requests_ref = db.collection('requests')
                recent_requests = requests_ref.order_by('createdAt', direction='DESCENDING').limit(10).stream()
                
                for req in recent_requests:
                    req_data = req.to_dict() or {}
                    activities.append({
                        'type': 'request',
                        'action': req_data.get('kind', 'unknown'),
                        'employeeId': req_data.get('employeeId', ''),
                        'supervisor': req_data.get('supervisor', ''),
                        'status': req_data.get('status', 'active'),
                        'timestamp': req_data.get('createdAt').isoformat() if req_data.get('createdAt') else None,
                        'details': f"طلب {req_data.get('kind', '')} للموظف {req_data.get('employeeId', '')}"
                    })
            except ResourceExhausted:
                # إذا تجاوزت الحصة نعيد نتيجة فارغة بدل انهيار العامل
                activities = []
        
        # ترتيب حسب التاريخ
        activities.sort(key=lambda x: x.get('timestamp', ''), reverse=True)
        result = activities[:10]
        cache.set(cache_key, result, timeout=30)
        return jsonify(result)
        
    except Exception as e:
        print(f"❌ خطأ في جلب الأنشطة: {str(e)}")
        return jsonify({"error": "خطأ في جلب الأنشطة"}), 500

@app.route("/api/employees/search", methods=["POST"])
@token_required
def search_employees_advanced(current_user):
    """بحث متقدم في الموظفين"""
    try:
        from firebase_config import get_all_employees
        
        data = request.get_json()
        query = data.get('query', '').lower()
        department = data.get('department', '')
        active_only = data.get('active_only', False)
        
        employees = get_all_employees()
        
        # تطبيق الفلاتر
        filtered_employees = employees
        
        # فلتر النص
        if query:
            filtered_employees = [
                e for e in filtered_employees
                if query in e.get('name', '').lower() 
                or query in str(e.get('employee_id', '')).lower()
                or query in e.get('department', '').lower()
            ]
        
        # فلتر القسم
        if department:
            filtered_employees = [
                e for e in filtered_employees
                if e.get('department', '') == department
            ]
        
        # فلتر النشطين فقط
        if active_only:
            filtered_employees = [
                e for e in filtered_employees
                if e.get('active', True)
            ]
        
        return jsonify({
            "total": len(filtered_employees),
            "employees": filtered_employees
        })
        
    except Exception as e:
        print(f"❌ خطأ في البحث: {str(e)}")
        return jsonify({"error": "خطأ في البحث"}), 500

@app.route("/api/employees/departments", methods=["GET"])
@token_required
def get_departments(current_user):
    """جلب قائمة الأقسام المتاحة"""
    try:
        from firebase_config import get_all_employees
        
        employees = get_all_employees()
        departments = set()
        
        for emp in employees:
            dept = emp.get('department', '')
            if dept:
                departments.add(dept)
        
        return jsonify({
            "departments": sorted(list(departments))
        })
        
    except Exception as e:
        print(f"❌ خطأ في جلب الأقسام: {str(e)}")
        return jsonify({"error": "خطأ في جلب الأقسام"}), 500


@app.route("/api/employees/upload-excel", methods=["POST"])
@token_required
def upload_employees_excel(current_user):
    """رفع ومعالجة ملف Excel لإضافة/تحديث الموظفين"""
    try:
        from firebase_config import db, update_employee, create_employee
        import openpyxl
        
        print(f"📤 استقبال طلب رفع ملف Excel للموظفين من {current_user}")
        
        # Check if file exists in request
        if 'file' not in request.files:
            return jsonify({"error": "لم يتم إرفاق ملف"}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({"error": "لم يتم اختيار ملف"}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({"error": "يجب أن يكون الملف بصيغة Excel (.xlsx أو .xls)"}), 400
        
        # Read Excel file in read_only mode to reduce memory usage
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        
        # Get headers from first row
        headers_row = [cell.value for cell in ws[1]]
        
        # Expected headers mapping (Arabic to English field names)
        field_mapping = {
            'رقم الموظف': 'employee_id',
            'الاسم': 'name',
            'المسمى الوظيفي': 'job_title',
            'الإدارة': 'department',
            'المدينة': 'city',
            'الدولة': 'country',
            'البريد الإلكتروني': 'email',
            'الهاتف': 'phone',
            'تاريخ البداية': 'start_date'
        }
        
        # Find column indices
        col_indices = {}
        for idx, header in enumerate(headers_row):
            if header and header in field_mapping:
                col_indices[field_mapping[header]] = idx
        
        print(f"📊 الأعمدة المتاحة: {list(col_indices.keys())}")
        
        # Validate required fields
        required_fields = ['employee_id', 'name']
        missing_fields = [f for f in required_fields if f not in col_indices]
        
        if missing_fields:
            missing_ar = [k for k, v in field_mapping.items() if v in missing_fields]
            return jsonify({"error": f"الأعمدة المطلوبة غير موجودة: {', '.join(missing_ar)}"}), 400
        
        # SMART OPTIMIZATION: Fetch all existing employee IDs once (single query)
        print(f"🔍 جلب قائمة الموظفين الموجودين...")
        existing_employee_ids = set()
        existing_employees_data = {}  # للمقارنة مع البيانات الجديدة
        try:
            # Get all employees with their employee_id field (not document ID)
            employees_ref = db.collection('employees')
            docs = employees_ref.stream()
            for doc in docs:
                doc_data = doc.to_dict()
                emp_id = doc_data.get('employee_id') or doc_data.get('id') or doc.id
                if emp_id:
                    existing_employee_ids.add(emp_id)
                    existing_employees_data[emp_id] = {
                        'doc_id': doc.id,  # معرف الوثيقة للتحديث
                        'data': doc_data   # البيانات الحالية للمقارنة
                    }
            print(f"✅ تم جلب {len(existing_employee_ids)} موظف موجود")
        except Exception as e:
            print(f"⚠️ خطأ في جلب الموظفين الموجودين: {e}")
            # Continue anyway - will treat all as new
        
        # Process rows with batch writes
        added = 0
        updated = 0
        skipped = 0
        errors = []
        processed = 0
        
        # Firestore batch for efficient writes (up to 500 operations)
        batch = db.batch()
        batch_operations = 0
        BATCH_SIZE = 100  # Commit every 100 operations
        
        print(f"📊 بدء معالجة ملف Excel بطريقة ذكية (تقليل استدعاءات Firestore)")
        
        # Stream process rows without loading all in memory
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Extract employee_id first to check if row is valid
            employee_id_col = col_indices.get('employee_id')
            if employee_id_col is None or employee_id_col >= len(row):
                skipped += 1
                continue
                
            employee_id = row[employee_id_col]
            # Check if employee_id is not None, not empty, and not just whitespace
            if employee_id is None or not str(employee_id).strip() or str(employee_id).strip() == 'None':
                skipped += 1
                continue
                
            # Also check if name exists
            name_col = col_indices.get('name')
            if name_col is None or name_col >= len(row):
                skipped += 1
                continue
                
            name = row[name_col]
            if name is None or not str(name).strip() or str(name).strip() == 'None':
                skipped += 1
                continue
            try:
                # Extract data based on column indices
                emp_data = {}
                
                for field, col_idx in col_indices.items():
                    value = row[col_idx] if col_idx < len(row) else None
                    if value is not None and value != '':
                        emp_data[field] = str(value).strip()
                
                employee_id = str(emp_data['employee_id']).strip()
                
                # CRITICAL: Always use employee_id as document ID to ensure uniqueness
                emp_ref = db.collection('employees').document(employee_id)
                
                # Check against in-memory set (instant - no network call!)
                if employee_id in existing_employee_ids:
                    # Get existing employee data for comparison
                    existing_emp = existing_employees_data[employee_id]
                    existing_data = existing_emp['data']
                    
                    # Compare data to see if update is needed
                    needs_update = False
                    update_data = {}
                    
                    # Compare each field
                    for field, new_value in emp_data.items():
                        if field == 'employee_id':
                            continue  # Skip employee_id - it's the key
                        
                        old_value = existing_data.get(field, '')
                        if str(old_value).strip() != str(new_value).strip():
                            update_data[field] = new_value
                            needs_update = True
                    
                    if needs_update:
                        update_data['updated_at'] = datetime.now().isoformat()
                        
                        # Add to batch - using employee_id as document ID
                        batch.update(emp_ref, update_data)
                        batch_operations += 1
                        updated += 1
                        
                        # Show what changed
                        changed_fields = list(update_data.keys())
                        changed_fields.remove('updated_at')
                        print(f"✅ تحديث: {employee_id} - {emp_data.get('name')} (تغيير: {', '.join(changed_fields)})")
                    else:
                        skipped += 1
                        print(f"⏭️ تجاهل: {employee_id} - {emp_data.get('name')} (لا توجد تغييرات)")
                else:
                    # Create new employee using batch
                    emp_data['id'] = employee_id  # Keep for backward compatibility
                    emp_data['employee_id'] = employee_id  # Ensure this field exists
                    emp_data['active'] = True
                    emp_data['created_at'] = datetime.now().isoformat()
                    emp_data['updated_at'] = datetime.now().isoformat()
                    emp_data['status'] = 'active'
                    
                    # CRITICAL: Use employee_id as document ID - this prevents duplicates
                    batch.set(emp_ref, emp_data)
                    batch_operations += 1
                    added += 1
                    print(f"➕ إضافة: {employee_id} - {emp_data.get('name')}")
                    
                    # Add to set for subsequent checks in same upload
                    existing_employee_ids.add(employee_id)
                    existing_employees_data[employee_id] = {
                        'doc_id': employee_id,  # document ID = employee_id
                        'data': emp_data
                    }
                
                processed += 1
                
                # Commit batch every BATCH_SIZE operations to avoid timeout
                if batch_operations >= BATCH_SIZE:
                    print(f"💾 حفظ دفعة من {batch_operations} عملية...")
                    batch.commit()
                    batch = db.batch()  # Create new batch
                    batch_operations = 0
                
                # Progress logging every 50 rows
                if processed % 50 == 0:
                    print(f"📊 تم معالجة {processed} صف...")
                    
            except Exception as row_error:
                error_msg = f"صف {row_idx}: {str(row_error)}"
                errors.append(error_msg)
                print(f"❌ {error_msg}")
                continue
        
        # Commit any remaining operations in final batch
        if batch_operations > 0:
            print(f"💾 حفظ الدفعة الأخيرة ({batch_operations} عملية)...")
            batch.commit()
        
        # Close workbook to free memory
        wb.close()
        
        result = {
            "success": True,
            "added": added,
            "updated": updated,
            "skipped": skipped,
            "total": added + updated,
            "errors": errors,
            "total_processed": processed
        }
        
        print(f"✅ اكتمل رفع الملف: {added} إضافة، {updated} تحديث، {skipped} تجاهل من أصل {processed} صف معالج")
        
        return jsonify(result)
        
    except Exception as e:
        print(f"❌ خطأ في رفع ملف Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"خطأ في معالجة الملف: {str(e)}"}), 500


# === نقاط النهاية العامة ===

@app.route("/api/health", methods=["GET"])
def health():
    """فحص صحة الخادم"""
    return jsonify({
        "status": "healthy",
        "firebase": firebase_initialized,
        "timestamp": datetime.utcnow().isoformat()
    })

@app.route("/", methods=["GET"])
def root():
    """الصفحة الرئيسية"""
    return jsonify({
        "message": "PreStaff API Server with Firebase",
        "version": "2.0.0",
        "firebase_enabled": firebase_initialized
    })

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
